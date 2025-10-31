#!/usr/bin/env python3
import os
import json
import csv
import asyncio
import logging
from datetime import datetime, timedelta, date, time
from typing import Dict, Optional, List, Tuple
import re
import io
import unicodedata

import pytz
import asyncpg
from telegram import Update
from telegram.constants import ChatType, ParseMode
from telegram.ext import Application, ContextTypes, MessageHandler, filters

# =============================
# LOGGING
# =============================
logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO,
)
log = logging.getLogger("bot")

# =============================
# CONFIG
# =============================
def get_env_variable(var_name: str) -> str:
    value = os.getenv(var_name)
    if not value:
        raise RuntimeError(
            f"Missing required environment variable '{var_name}'. "
            "Please set it in your hosting environment."
        )
    return value

BOT_TOKEN = get_env_variable("BOT_TOKEN")
DATABASE_URL = get_env_variable("DATABASE_URL")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "excelmerge")  # telegram username (without @)
WA_DAILY_LIMIT = int(os.getenv("WA_DAILY_LIMIT", "2"))        # max sends per number per logical day
REMINDER_DELAY_MINUTES = int(os.getenv("REMINDER_DELAY_MINUTES", "30")) # Delay for reminders
USER_WHATSAPP_LIMIT = int(os.getenv("USER_WHATSAPP_LIMIT", "10"))
USERNAME_THRESHOLD_FOR_BONUS = int(os.getenv("USERNAME_THRESHOLD_FOR_BONUS", "25"))
REQUEST_GROUP_ID = int(os.getenv("REQUEST_GROUP_ID", "-1002438185636")) # Group for 'i need ...' commands
CLEARING_GROUP_ID = int(os.getenv("CLEARING_GROUP_ID", "-1002624324856")) # Group for auto-clearing pendings
CONFIRMATION_GROUP_ID = int(os.getenv("CONFIRMATION_GROUP_ID", "-1002694540582"))
DETAIL_GROUP_ID = int(os.getenv("DETAIL_GROUP_ID", "-1002598927727")) # Group for 'my detail' reports
# --- TARGET GROUP FOR FORWARDED MESSAGES ---
FORWARD_GROUP_ID = int(os.getenv("FORWARD_GROUP_ID", "-1003109226804")) # Target for cleared messages
# ------------------------------------------
PERFORMANCE_GROUP_IDS = {
    -1002670785417, -1002659012767, -1002790753092, -1002520117752
}

# NEW: Daily limit for submissions per country
COUNTRY_DAILY_LIMIT = int(os.getenv("COUNTRY_DAILY_LIMIT", "15"))


# Whitelist of allowed countries (lowercase for case-insensitive matching)
ALLOWED_COUNTRIES = {
    'morocco', 'panama', 'saudi arabia', 'united arab emirates', 'uae',
    'oman', 'jordan', 'italy', 'germany', 'indonesia', 'colombia',
    'bulgaria', 'brazil', 'spain', 'belgium', 'south africa',
    'philippines', 'portugal', 'netherlands', 'poland', 'ghana', 'dominican republic',
    'qatar', 'france', 'switzerland', 'argentina', 'costa rica', 'pakistan', 'kuwait', 'yemen', 'bahrain', 'malaysia',
    'canada','mauritania'
}

TIMEZONE = pytz.timezone("Asia/Phnom_Penh")

# =============================
# DATABASE SETUP & HELPERS
# =============================
db_lock = asyncio.Lock()
DB_POOL: Optional[asyncpg.Pool] = None

async def get_db_pool() -> asyncpg.Pool:
    global DB_POOL
    if DB_POOL is None or DB_POOL.is_closing():
        try:
            DB_POOL = await asyncpg.create_pool(
                dsn=DATABASE_URL,
                max_inactive_connection_lifetime=60,
                min_size=1,
                max_size=10
            )
            if DB_POOL is None:
                raise ConnectionError("Database pool initialization failed, create_pool returned None.")
            log.info("Database connection pool established.")
        except Exception as e:
            log.error(f"Could not create database connection pool: {e}")
            raise
    return DB_POOL

async def close_db_pool():
    global DB_POOL
    if DB_POOL and not DB_POOL.is_closing():
        log.info("Closing database connection pool.")
        await DB_POOL.close()
        DB_POOL = None
        log.info("Database connection pool closed.")


async def setup_database():
    log.info("Setting up database schema...")
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS kv_storage (
                key TEXT PRIMARY KEY,
                data JSONB NOT NULL,
                updated_at TIMESTAMPTZ DEFAULT NOW()
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id SERIAL PRIMARY KEY,
                ts_local TIMESTAMPTZ NOT NULL,
                chat_id BIGINT,
                message_id BIGINT,
                user_id BIGINT,
                user_first TEXT,
                user_username TEXT,
                kind TEXT,
                action TEXT,
                value TEXT,
                owner TEXT
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS wa_daily_usage (
                day DATE NOT NULL,
                number_norm TEXT NOT NULL,
                sent_count INTEGER NOT NULL DEFAULT 0,
                last_sent TIMESTAMPTZ,
                PRIMARY KEY (day, number_norm)
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS user_daily_activity (
                day DATE NOT NULL,
                user_id BIGINT NOT NULL,
                username_requests INTEGER DEFAULT 0,
                whatsapp_requests INTEGER DEFAULT 0,
                PRIMARY KEY (day, user_id)
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS whatsapp_bans (
                user_id BIGINT PRIMARY KEY
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS whitelisted_users (
                user_id BIGINT PRIMARY KEY
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS user_daily_country_counts (
                day DATE NOT NULL,
                user_id BIGINT NOT NULL,
                country TEXT NOT NULL,
                count INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (day, user_id, country)
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS user_daily_confirmations (
                day DATE NOT NULL,
                user_id BIGINT NOT NULL,
                confirm_count INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (day, user_id)
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS owner_daily_performance (
                day DATE NOT NULL,
                owner_name TEXT NOT NULL,
                telegram_count INTEGER NOT NULL DEFAULT 0,
                whatsapp_count INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (day, owner_name)
            );
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS admins (
                username TEXT PRIMARY KEY,
                permissions JSONB NOT NULL
            );
        """)
        # This table is for MANUAL, PERMANENT bans
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS user_country_bans (
                user_id BIGINT NOT NULL,
                country TEXT NOT NULL,
                PRIMARY KEY (user_id, country)
            );
        """)
    log.info("Database schema is ready.")


# =============================
# STATE (DB-backed)
# =============================
BASE_STATE = {
    "user_names": {},
    "rr": {
        "username_owner_idx": 0, "username_entry_idx": {},
        "wa_owner_idx": 0, "wa_entry_idx": {},
    },
    "issued": {"username": {}, "whatsapp": {}, "app_id": {}},
    "priority_queue": {
        "active": False,
        "owner": None,
        "remaining": 0,
        "stop_after": False,
        "saved_rr_indices": {}
    },
    "whatsapp_temp_bans": {},
    "whatsapp_last_request_ts": {},
    "username_last_request_ts": {},
    "whatsapp_offense_count": {}
}
state: Dict = {k: (v.copy() if isinstance(v, dict) else v) for k, v in BASE_STATE.items()}
WHATSAPP_BANNED_USERS: set[int] = set()
WHITELISTED_USERS: set[int] = set()
ADMIN_PERMISSIONS: Dict[str, List[str]] = {}
USER_COUNTRY_BANS: Dict[int, set[str]] = {}

# =============================
# PERMISSIONS & ADMINS
# =============================
COMMAND_PERMISSIONS = {
    'add owner', 'delete owner', 'add username', 'delete username', 'add whatsapp', 'delete whatsapp',
    'stop open', 'take customer', 'ban whatsapp', 'unban whatsapp', 'report',
    'owner report', 'performance', 'remind user', 'clear pending',
    'list owners', 'list disabled', 'list owner', 'detail user', 'list banned', 'list admins',
    'data today', 'list enabled', 'add user', 'delete user',
    'ban country', 'unban country', 'list country bans', 'user performance', 'user stats',
    'inventory', 'request stats'
}

async def load_admins():
    global ADMIN_PERMISSIONS
    ADMIN_PERMISSIONS = {}
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("SELECT username, permissions FROM admins")
            for row in rows:
                ADMIN_PERMISSIONS[row['username']] = json.loads(row['permissions'])
        log.info(f"Loaded {len(ADMIN_PERMISSIONS)} admins from database.")
    except Exception as e:
        log.error(f"Failed to load admins from DB: %s", e)

def _is_super_admin(user: Optional[Update.effective_user]) -> bool:
    if not user: return False
    return (user.username or "").lower() == ADMIN_USERNAME.lower()

def _is_admin(user: Optional[Update.effective_user]) -> bool:
    if not user or not user.username: return False
    return _is_super_admin(user) or _norm_owner_name(user.username) in ADMIN_PERMISSIONS

def _has_permission(user: Optional[Update.effective_user], permission: str) -> bool:
    if not user or not user.username: return False
    if _is_super_admin(user): return True

    user_permissions = ADMIN_PERMISSIONS.get(_norm_owner_name(user.username), [])
    return permission in user_permissions

async def load_whatsapp_bans():
    global WHATSAPP_BANNED_USERS
    WHATSAPP_BANNED_USERS = set()
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("SELECT user_id FROM whatsapp_bans")
            for row in rows:
                WHATSAPP_BANNED_USERS.add(row['user_id'])
        log.info(f"Loaded {len(WHATSAPP_BANNED_USERS)} WhatsApp bans from database.")
    except Exception as e:
        log.error(f"Failed to load WhatsApp bans from DB: %s", e)

async def load_user_country_bans():
    global USER_COUNTRY_BANS
    USER_COUNTRY_BANS = {}
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("SELECT user_id, country FROM user_country_bans")
            for row in rows:
                user_id = row['user_id']
                country = row['country'].lower()
                if user_id not in USER_COUNTRY_BANS:
                    USER_COUNTRY_BANS[user_id] = set()
                USER_COUNTRY_BANS[user_id].add(country)
        log.info(f"Loaded {sum(len(c) for c in USER_COUNTRY_BANS.values())} user-country (manual) bans from database.")
    except Exception as e:
        log.error(f"Failed to load user-country bans from DB: %s", e)

async def load_whitelisted_users():
    global WHITELISTED_USERS
    WHITELISTED_USERS = set()
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("SELECT user_id FROM whitelisted_users")
            for row in rows:
                WHITELISTED_USERS.add(row['user_id'])
        log.info(f"Loaded {len(WHITELISTED_USERS)} whitelisted users from database.")
    except Exception as e:
        log.error(f"Failed to load whitelisted users from DB: %s", e)

async def load_state():
    global state
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            result = await conn.fetchval("SELECT data FROM kv_storage WHERE key = 'state'")
            if result:
                loaded = json.loads(result)
                temp_state = {k: (v.copy() if isinstance(v, dict) else v) for k, v in BASE_STATE.items()}
                for key, value in loaded.items():
                    if isinstance(value, dict) and key in temp_state:
                        temp_state[key].update(value)
                    else:
                        temp_state[key] = value
                state = temp_state
                log.info("Bot state loaded from database.")
            else:
                state = {k: (v.copy() if isinstance(v, dict) else v) for k, v in BASE_STATE.items()}
                await save_state()
    except Exception as e:
        log.warning(f"Failed to load state from DB: %s. Using default state.", e)
        state = {k: (v.copy() if isinstance(v, dict) else v) for k, v in BASE_STATE.items()}

    state.setdefault("rr", {}).setdefault("username_entry_idx", {})
    state["rr"].setdefault("wa_entry_idx", {})
    state.setdefault("issued", {}).setdefault("username", {})
    state["issued"].setdefault("whatsapp", {})
    state["issued"].setdefault("app_id", {})
    state.setdefault("priority_queue", BASE_STATE["priority_queue"])
    state.setdefault("whatsapp_temp_bans", {})
    state.setdefault("whatsapp_last_request_ts", {})
    state.setdefault("username_last_request_ts", {})
    state.setdefault("whatsapp_offense_count", {})


async def save_state():
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO kv_storage (key, data)
                VALUES ('state', $1)
                ON CONFLICT (key) DO UPDATE
                SET data = EXCLUDED.data, updated_at = NOW();
            """, json.dumps(state))
    except Exception as e:
        log.warning("Failed to save state to DB: %s", e)


async def _migrate_state_if_needed():
    log.info("Checking state structure for migration...")
    state_was_changed = False
    issued = state.setdefault("issued", {})
    for kind in ("username", "whatsapp", "app_id"):
        bucket = issued.setdefault(kind, {})
        for user_id, item_or_list in bucket.items():
            if isinstance(item_or_list, dict):
                bucket[user_id] = [item_or_list]
                state_was_changed = True
                log.info(f"Migrated user {user_id}'s '{kind}' data to new list format.")

    if state_was_changed:
        log.info("State structure was migrated. Saving new format to database.")
        await save_state()
    else:
        log.info("State structure is already up-to-date.")

# =============================
# OWNER DIRECTORY (DB-backed)
# =============================
OWNER_DATA: List[dict] = []
HANDLE_INDEX: Dict[str, List[dict]] = {}
PHONE_INDEX: Dict[str, dict] = {}
USERNAME_POOL: List[Dict] = []
WHATSAPP_POOL: List[Dict] = []

def _norm_handle(h: str) -> str: return re.sub(r"^@", "", (h or "").strip().lower())
def _norm_phone(p: str) -> str: return re.sub(r"\D+", "", (p or ""))
def _normalize_app_id(app_id: str) -> str:
    """Removes leading '@', normalizes unicode characters, removes non-alphanumeric, and converts to lowercase."""
    if not app_id:
        return ""
    normalized_str = unicodedata.normalize('NFKC', app_id)
    return re.sub(r'[^a-zA-Z0-9]', '', normalized_str).lower()


def _norm_owner_name(s: str) -> str:
    s = (s or "").strip()
    if s.startswith("@"): s = s[1:]
    return s.lower()

def _is_owner(user: Optional[Update.effective_user]) -> bool:
    if not user or not user.username:
        return False
    norm_username = _norm_owner_name(user.username)
    return any(_norm_owner_name(g.get("owner", "")) == norm_username for g in OWNER_DATA)

def _owner_is_paused(group: dict) -> bool:
    if group.get("disabled") or group.get("active") is False: return True
    until = group.get("disabled_until")
    if until:
        try:
            dt = datetime.fromisoformat(until)
            if dt.tzinfo is None: dt = TIMEZONE.localize(dt)
            return datetime.now(tz=TIMEZONE) < dt
        except Exception:
            return False
    return False

def _ensure_owner_shape(g: dict) -> dict:
    g.setdefault("owner", "")
    g.setdefault("disabled", False)
    g.setdefault("managed_by", None)
    g.setdefault("entries", [])
    g.setdefault("whatsapp", [])

    norm_entries = []
    for e in g.get("entries", []):
        if isinstance(e, dict):
            e_copy = e.copy()
            e_copy.setdefault("telegram", "")
            e_copy.setdefault("phone", "")
            e_copy.setdefault("disabled", False)
            e_copy.setdefault("managed_by", None)
            norm_entries.append(e_copy)
    g["entries"] = norm_entries

    norm_wa = []
    for w in g.get("whatsapp", []):
        entry = {}
        if isinstance(w, dict):
            entry = w.copy()
            entry.setdefault("number", w.get("number") or w.get("phone") or "")
        elif isinstance(w, str) and w.strip():
            entry["number"] = w

        if (entry.get("number") or "").strip():
            entry.setdefault("disabled", False)
            entry.setdefault("managed_by", None)
            norm_wa.append(entry)

    g["whatsapp"] = norm_wa
    return g

async def save_owner_directory():
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO kv_storage (key, data)
                VALUES ('owners', $1)
                ON CONFLICT (key) DO UPDATE
                SET data = EXCLUDED.data, updated_at = NOW();
            """, json.dumps(OWNER_DATA))
    except Exception as e:
        log.error("Failed to save owner directory to DB: %s", e)

async def load_owner_directory():
    global OWNER_DATA, HANDLE_INDEX, PHONE_INDEX, USERNAME_POOL, WHATSAPP_POOL
    OWNER_DATA, HANDLE_INDEX, PHONE_INDEX = [], {}, {}
    USERNAME_POOL, WHATSAPP_POOL = [], []

    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            result = await conn.fetchval("SELECT data FROM kv_storage WHERE key = 'owners'")
            OWNER_DATA = (json.loads(result) or []) if result else []
    except Exception as e:
        log.error("Failed to load owners from DB: %s", e)
        OWNER_DATA = []

    OWNER_DATA = [_ensure_owner_shape(dict(g)) for g in OWNER_DATA]

    for group in OWNER_DATA:
        owner = _norm_owner_name(group.get("owner") or "")
        if not owner or _owner_is_paused(group): continue
        usernames: List[str] = []
        for entry in group.get("entries", []):
            if entry.get("disabled"): continue
            tel = (entry.get("telegram") or "").strip()
            ph  = (entry.get("phone") or "").strip()
            if tel:
                handle_shown = tel if tel.startswith("@") else f"@{tel}"
                usernames.append(handle_shown)
                HANDLE_INDEX.setdefault(_norm_handle(tel), []).append(
                    {"owner": owner, "phone": ph, "telegram": tel, "channel": "telegram"})
            if ph:
                PHONE_INDEX[_norm_phone(ph)] = {"owner": owner, "phone": ph, "telegram": tel, "channel": "telegram"}
        if usernames: USERNAME_POOL.append({"owner": owner, "usernames": usernames})
        numbers: List[str] = []
        for w in group.get("whatsapp", []):
            if w.get("disabled"): continue
            num = (w.get("number") or "").strip()
            if num:
                numbers.append(num)
                PHONE_INDEX[_norm_phone(num)] = {"owner": owner, "phone": num, "telegram": None, "channel": "whatsapp"}
        if numbers: WHATSAPP_POOL.append({"owner": owner, "numbers": numbers})

    log.info("[owner_directory] owners(active): usernames=%d, whatsapp=%d | handles=%d, phones=%d",
             len(USERNAME_POOL), len(WHATSAPP_POOL), len(HANDLE_INDEX), len(PHONE_INDEX))

# =============================
# QUOTA & USER ACTIVITY
# =============================
def _logical_day_today() -> date:
    now = datetime.now(TIMEZONE)
    return (now - timedelta(hours=5, minutes=30)).date()

async def _get_user_activity(user_id: int) -> Tuple[int, int]:
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT username_requests, whatsapp_requests FROM user_daily_activity WHERE day=$1 AND user_id=$2",
                _logical_day_today(), user_id
            )
            return (row['username_requests'], row['whatsapp_requests']) if row else (0, 0)
    except Exception as e:
        log.warning(f"User activity read failed for {user_id}: {e}")
        return (0, 0)

async def _increment_user_activity(user_id: int, kind: str):
    # This function correctly only increments the counts and is called only on a successful request.
    # No changes are needed here as it aligns with the user's request.
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            if kind == "username":
                await conn.execute("""
                    INSERT INTO user_daily_activity (day, user_id, username_requests)
                    VALUES ($1, $2, 1)
                    ON CONFLICT (day, user_id) DO UPDATE
                    SET username_requests = user_daily_activity.username_requests + 1;
                """, _logical_day_today(), user_id)
            elif kind == "whatsapp":
                await conn.execute("""
                    INSERT INTO user_daily_activity (day, user_id, whatsapp_requests)
                    VALUES ($1, $2, 1)
                    ON CONFLICT (day, user_id) DO UPDATE
                    SET whatsapp_requests = user_daily_activity.whatsapp_requests + 1;
                """, _logical_day_today(), user_id)
    except Exception as e:
        log.warning(f"User activity write failed for {user_id}: {e}")

async def _increment_user_country_count(user_id: int, country: str):
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO user_daily_country_counts (day, user_id, country, count)
                VALUES ($1, $2, $3, 1)
                ON CONFLICT (day, user_id, country) DO UPDATE
                SET count = user_daily_country_counts.count + 1;
            """, _logical_day_today(), user_id, country)
    except Exception as e:
        log.warning(f"User country count write failed for {user_id} and country {country}: {e}")

# NEW: Helper function to get the current count for a single country
async def _get_user_country_count(user_id: int, country: str) -> int:
    """Fetches a user's submission count for a specific country on the current logical day."""
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            count = await conn.fetchval(
                "SELECT count FROM user_daily_country_counts WHERE day=$1 AND user_id=$2 AND country=$3",
                _logical_day_today(), user_id, country
            )
            return int(count) if count is not None else 0
    except Exception as e:
        log.warning(f"User country count read failed for {user_id} and country {country}: {e}")
        return 0


async def _increment_user_confirmation_count(user_id: int):
    """Increments the successful confirmation count for a user on the current logical day."""
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO user_daily_confirmations (day, user_id, confirm_count)
                VALUES ($1, $2, 1)
                ON CONFLICT (day, user_id) DO UPDATE
                SET confirm_count = user_daily_confirmations.confirm_count + 1;
            """, _logical_day_today(), user_id)
    except Exception as e:
        log.warning(f"User confirmation count write failed for {user_id}: {e}")

async def _increment_owner_performance(owner_name: str, kind: Optional[str]):
    """Increments telegram or whatsapp count for an owner on the current day."""
    if not owner_name:
        return
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            if kind == 'username':
                await conn.execute("""
                    INSERT INTO owner_daily_performance (day, owner_name, telegram_count)
                    VALUES ($1, $2, 1)
                    ON CONFLICT (day, owner_name) DO UPDATE
                    SET telegram_count = owner_daily_performance.telegram_count + 1;
                """, _logical_day_today(), owner_name)
            elif kind == 'whatsapp':
                await conn.execute("""
                    INSERT INTO owner_daily_performance (day, owner_name, whatsapp_count)
                    VALUES ($1, $2, 1)
                    ON CONFLICT (day, owner_name) DO UPDATE
                    SET whatsapp_count = owner_daily_performance.whatsapp_count + 1;
                """, _logical_day_today(), owner_name)
            # For kind == 'app_id' or None, there is no performance metric to update.
    except Exception as e:
        log.warning(f"Owner performance write failed for {owner_name}: {e}")


async def _wa_get_count(number_norm: str, day: date) -> int:
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            count = await conn.fetchval(
                "SELECT sent_count FROM wa_daily_usage WHERE day=$1 AND number_norm=$2",
                day, number_norm
            )
            return int(count) if count is not None else 0
    except Exception as e:
        log.warning("Quota read failed: %s", e)
        return 0

async def _wa_inc_count(number_norm: str, day: date):
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO wa_daily_usage (day, number_norm, sent_count, last_sent)
                VALUES ($1, $2, 1, $3)
                ON CONFLICT (day, number_norm)
                DO UPDATE SET sent_count = wa_daily_usage.sent_count + 1,
                              last_sent = EXCLUDED.last_sent
            """, day, number_norm, datetime.now(TIMEZONE))
    except Exception as e:
        log.warning("Quota write failed: %s", e)

async def _wa_quota_reached(number_raw: str) -> bool:
    n = _norm_phone(number_raw)
    cnt = await _wa_get_count(n, _logical_day_today())
    return cnt >= WA_DAILY_LIMIT

# =============================
# Rotation-preserving rebuild & Round-robin helpers
# =============================
def _owner_list_from_pool(pool) -> List[str]: return [blk["owner"] for blk in pool]
def _preserve_owner_pointer(old_list: List[str], new_list: List[str], old_idx: int) -> int:
    if not new_list: return 0
    old_list = list(old_list or [])
    if not old_list: return 0
    if old_idx >= len(old_list): old_idx = 0 # Bounds check
    start_owner = old_list[old_idx]
    if start_owner in new_list: return new_list.index(start_owner)
    n = len(old_list)
    for step in range(1, n + 1):
        cand = old_list[(old_idx + step) % n]
        if cand in new_list: return new_list.index(cand)
    return 0

def _preserve_entry_indices(rr_map: Dict[str, int], new_pool: List[Dict], list_key: str):
    valid_owners = {blk["owner"]: len(blk.get(list_key, []) or []) for blk in new_pool}
    for owner in list(rr_map.keys()):
        if owner not in valid_owners: rr_map.pop(owner, None)
    for owner, sz in valid_owners.items():
        if sz <= 0: rr_map.pop(owner, None)
        else: rr_map[owner] = rr_map.get(owner, 0) % sz

async def _rebuild_pools_preserving_rotation():
    old_user_owner_list = _owner_list_from_pool(USERNAME_POOL)
    old_wa_owner_list   = _owner_list_from_pool(WHATSAPP_POOL)

    rr = state.setdefault("rr", {})
    old_user_owner_idx = rr.get("username_owner_idx", 0)
    old_wa_owner_idx   = rr.get("wa_owner_idx", 0)
    old_user_entry_idx = dict(rr.get("username_entry_idx", {}))
    old_wa_entry_idx   = dict(rr.get("wa_entry_idx", {}))

    await save_owner_directory()
    await load_owner_directory()

    new_user_owner_list = _owner_list_from_pool(USERNAME_POOL)
    new_wa_owner_list   = _owner_list_from_pool(WHATSAPP_POOL)

    rr["username_owner_idx"] = _preserve_owner_pointer(
        old_user_owner_list, new_user_owner_list, old_user_owner_idx
    )
    rr["wa_owner_idx"] = _preserve_owner_pointer(
        old_wa_owner_list, new_wa_owner_list, old_wa_owner_idx
    )

    rr.setdefault("username_entry_idx", old_user_entry_idx)
    rr.setdefault("wa_entry_idx", old_wa_entry_idx)
    _preserve_entry_indices(rr["username_entry_idx"], USERNAME_POOL, "usernames")
    _preserve_entry_indices(rr["wa_entry_idx"], WHATSAPP_POOL, "numbers")

    await save_state()

async def _decrement_priority_and_end_if_needed():
    pq = state.get("priority_queue", {})
    if not pq.get("active"):
        return

    pq["remaining"] -= 1

    if pq["remaining"] <= 0:
        log.info(f"Priority queue for owner {pq['owner']} completed.")
        saved_indices = pq.get("saved_rr_indices", {})
        state["rr"]["username_owner_idx"] = saved_indices.get("username_owner_idx", 0)
        state["rr"]["wa_owner_idx"] = saved_indices.get("wa_owner_idx", 0)

        stop_after = pq.get("stop_after", False)
        owner_to_stop = pq.get("owner")

        state["priority_queue"] = BASE_STATE["priority_queue"]

        if stop_after and owner_to_stop:
            log.info(f"Auto-stopping owner {owner_to_stop} after priority queue completion.")
            owner_group = _find_owner_group(owner_to_stop)
            if owner_group:
                owner_group["disabled"] = True
            await _rebuild_pools_preserving_rotation()
        else:
            await save_state()
    else:
        await save_state()

async def _next_from_username_pool() -> Optional[Dict[str, str]]:
    pq = state.get("priority_queue", {})
    if pq.get("active"):
        priority_owner = pq.get("owner")
        for block in USERNAME_POOL:
            if block["owner"] == priority_owner:
                arr = block.get("usernames", [])
                if arr:
                    ei = state["rr"]["username_entry_idx"].get(priority_owner, 0) % len(arr)
                    result = {"owner": priority_owner, "username": arr[ei]}
                    state["rr"]["username_entry_idx"][priority_owner] = (ei + 1) % len(arr)
                    await _decrement_priority_and_end_if_needed()
                    return result
        log.warning(f"Priority owner {priority_owner} has no available usernames. Falling back to normal rotation for this request.")

    if not USERNAME_POOL: return None

    owner_idx = state['rr'].get("username_owner_idx", 0)
    if owner_idx >= len(USERNAME_POOL): owner_idx = 0

    # Start searching from the selected owner's index
    for i in range(len(USERNAME_POOL)):
        current_idx = (owner_idx + i) % len(USERNAME_POOL)
        block = USERNAME_POOL[current_idx]
        arr = block.get("usernames", [])
        if arr:
            ei = state["rr"]["username_entry_idx"].get(block["owner"], 0) % len(arr)
            result = {"owner": block["owner"], "username": arr[ei]}
            state["rr"]["username_entry_idx"][block["owner"]] = (ei + 1) % len(arr)
            state["rr"]["username_owner_idx"] = (current_idx + 1) % len(USERNAME_POOL)
            await save_state()
            return result
    return None


async def _next_from_whatsapp_pool() -> Optional[Dict[str, str]]:
    pq = state.get("priority_queue", {})
    if pq.get("active"):
        priority_owner = pq.get("owner")
        for block in WHATSAPP_POOL:
            if block["owner"] == priority_owner:
                numbers = block.get("numbers", []) or []
                if numbers:
                    start = state["rr"]["wa_entry_idx"].get(priority_owner, 0) % len(numbers)
                    for step in range(len(numbers)):
                        cand = numbers[(start + step) % len(numbers)]
                        if not await _wa_quota_reached(cand):
                            state["rr"]["wa_entry_idx"][priority_owner] = ((start + step) + 1) % len(numbers)
                            await _decrement_priority_and_end_if_needed()
                            return {"owner": priority_owner, "number": cand}
        log.warning(f"Priority owner {priority_owner} has no available WhatsApp numbers. Falling back to normal rotation for this request.")

    if not WHATSAPP_POOL: return None

    owner_idx = state['rr'].get("wa_owner_idx", 0)
    if owner_idx >= len(WHATSAPP_POOL): owner_idx = 0

    # Start searching from the selected owner's index
    for i in range(len(WHATSAPP_POOL)):
        current_idx = (owner_idx + i) % len(WHATSAPP_POOL)
        block = WHATSAPP_POOL[current_idx]
        owner = block["owner"]
        numbers = block.get("numbers", []) or []
        if numbers:
            start = state["rr"]["wa_entry_idx"].get(owner, 0) % len(numbers)
            for step in range(len(numbers)):
                cand = numbers[(start + step) % len(numbers)]
                if await _wa_quota_reached(cand):
                    continue
                state["rr"]["wa_entry_idx"][owner] = ((start + step) + 1) % len(numbers)
                state["rr"]["wa_owner_idx"] = (current_idx + 1) % len(WHATSAPP_POOL)
                await save_state()
                return {"owner": owner, "number": cand}
    return None

# =============================
# REGEXES & HELPERS
# =============================
WHO_USING_REGEX = re.compile(
    r"^\s*who(?:['\u2019']s| is)\s+using\s+(?:@?([A-Za-z0-9_\.]+)|(\+?\d[\d\s\-]{6,}\d))\s*$",
    re.IGNORECASE
)
NEED_USERNAME_RX = re.compile(r"^\s*i\s*need\s*(?:user\s*name|username)\s*$", re.IGNORECASE)
NEED_WHATSAPP_RX = re.compile(r"^\s*i\s*need\s*(?:id\s*)?whats?app\s*$", re.IGNORECASE)
APP_ID_RX = re.compile(r"\b(app|add|id)\b.*?\@([^\s]+)", re.IGNORECASE)
EXTRACT_USERNAMES_RX = re.compile(r'@([a-zA-Z0-9_]{4,})')
EXTRACT_PHONES_RX = re.compile(r'(\+?\d[\d\s\-()]{8,}\d)')


STOP_OPEN_RX          = re.compile(r"^\s*(stop|open)\s+(.+?)\s*$", re.IGNORECASE)
ADD_OWNER_RX          = re.compile(r"^\s*add\s+owner\s+@?(.+?)\s*$", re.IGNORECASE)
ADD_USERNAME_RX       = re.compile(r"^\s*add\s+username\s+@([A-Za-z0-9_]{3,})\s+to\s+@?(.+?)\s*$", re.IGNORECASE)
ADD_WHATSAPP_RX       = re.compile(r"^\s*add\s+whats?app\s+(\+?\d[\d\s\-]{6,}\d)\s+to\s+@?(.+?)\s*$", re.IGNORECASE)
DEL_OWNER_RX          = re.compile(r"^\s*delete\s+owner\s+@?(.+?)\s*$", re.IGNORECASE)
DEL_USERNAME_RX       = re.compile(r"^\s*delete\s+username\s+@([A-Za-z0-9_]{3,})\s*$", re.IGNORECASE)
DEL_WHATSAPP_RX       = re.compile(r"^\s*delete\s+whats?app\s+(\+?\d[\d\s\-]{6,}\d)\s*$", re.IGNORECASE)
LIST_OWNERS_RX        = re.compile(r"^\s*list\s+owners\s*$", re.IGNORECASE)
LIST_OWNER_DETAIL_RX  = re.compile(r"^\s*list\s+owner\s+@?([A-Za-z0-9_]{3,})\s*$", re.IGNORECASE)
LIST_DISABLED_RX      = re.compile(r"^\s*list\s+disabled\s*$", re.IGNORECASE)
SEND_REPORT_RX        = re.compile(r"^\s*(?:send\s+report|report)(?:\s+(yesterday|today|\d{4}-\d{2}-\d{2}))?\s*$", re.IGNORECASE)
PHONE_LIKE_RX         = re.compile(r"^\+?\d[\d\s\-]{6,}\d$")
LIST_OWNER_ALIAS_RX   = re.compile(r"^\s*list\s+@?([A-Za-z0-9_]{3,})\s*$", re.IGNORECASE)
REMIND_ALL_RX         = re.compile(r"^\s*remind\s+user\s*$", re.IGNORECASE)
TAKE_CUSTOMER_RX      = re.compile(r"^\s*take\s+(\d+)\s+customer(?:s)?\s+to\s+owner\s+@?(.+?)(?:\s+(and\s+stop))?\s*$", re.IGNORECASE)
CLEAR_PENDING_RX      = re.compile(r"^\s*clear\s+pending\s+(.+)\s*$", re.IGNORECASE)
BAN_WHATSAPP_RX       = re.compile(r"^\s*ban\s+whatsapp\s+@?(\S+)\s*$", re.IGNORECASE)
UNBAN_WHATSAPP_RX     = re.compile(r"^\s*unban\s+whatsapp\s+@?(\S+)\s*$", re.IGNORECASE)
LIST_BANNED_RX        = re.compile(r"^\s*list\s+banned\s*$", re.IGNORECASE)
OWNER_REPORT_RX       = re.compile(r"^\s*owner\s+report(?:\s+(yesterday|today|\d{4}-\d{2}-\d{2}))?\s*$", re.IGNORECASE)
COMMANDS_RX           = re.compile(r"^\s*commands\s*$", re.IGNORECASE)
MY_DETAIL_RX          = re.compile(r"^\s*my\s+detail\s*$", re.IGNORECASE)
DETAIL_USER_RX        = re.compile(r"^\s*detail\s+@?(\S+)\s*$", re.IGNORECASE)
MY_PERFORMANCE_RX     = re.compile(r"^\s*my\s+performance(?:\s+(yesterday|today|\d{4}-\d{2}-\d{2}))?\s*$", re.IGNORECASE)
PERFORMANCE_OWNER_RX  = re.compile(r"^\s*performance\s+@?(\S+?)(?:\s+(yesterday|today|\d{4}-\d{2}-\d{2}))?\s*$", re.IGNORECASE)
ADD_ADMIN_RX          = re.compile(r"^\s*add\s+admin\s+@?(\S+)\s*$", re.IGNORECASE)
DELETE_ADMIN_RX       = re.compile(r"^\s*delete\s+admin\s+@?(\S+)\s*$", re.IGNORECASE)
ALLOW_ADMIN_CMD_RX    = re.compile(r"^\s*allow\s+@?(\S+)\s+to\s+use\s+command\s+(.+)\s*$", re.IGNORECASE)
STOP_ALLOW_ADMIN_CMD_RX = re.compile(r"^\s*stop\s+allow\s+@?(\S+)\s+to\s+use\s+command\s+(.+)\s*$", re.IGNORECASE)
LIST_ADMINS_RX        = re.compile(r"^\s*list\s+admins\s*$", re.IGNORECASE)
DATA_TODAY_RX         = re.compile(r"^\s*data\s+today\s*$", re.IGNORECASE)
LIST_ENABLED_RX       = re.compile(r"^\s*list\s+enabled\s*$", re.IGNORECASE)
ADD_USER_RX           = re.compile(r"^\s*add\s+user\s+@?(\S+)\s*$", re.IGNORECASE)
DELETE_USER_RX        = re.compile(r"^\s*delete\s+user\s+@?(\S+)\s*$", re.IGNORECASE)
BAN_COUNTRY_RX        = re.compile(r"^\s*ban\s+country\s+([a-zA-Z\s]+)\s+for\s+@?(\S+)\s*$", re.IGNORECASE)
UNBAN_COUNTRY_RX      = re.compile(r"^\s*unban\s+country\s+([a-zA-Z\s]+)\s+for\s+@?(\S+)\s*$", re.IGNORECASE)
LIST_COUNTRY_BANS_RX  = re.compile(r"^\s*list\s+country\s+bans(?:\s+@?(\S+))?\s*$", re.IGNORECASE)
USER_PERFORMANCE_RX   = re.compile(r"^\s*user\s+performance(?:\s+(today|yesterday|\d{4}-\d{2}-\d{2}))?\s*$", re.IGNORECASE)
USER_STATS_RX         = re.compile(r"^\s*user\s+stats(?:\s+(today|yesterday|\d{4}-\d{2}-\d{2}))?\s*$", re.IGNORECASE)
INVENTORY_RX          = re.compile(r"^\s*inventory\s*$", re.IGNORECASE)
REQUEST_STATS_RX      = re.compile(r"^\s*request\s+stats\s*$", re.IGNORECASE)


def _looks_like_phone(s: str) -> bool:
    return bool(PHONE_LIKE_RX.fullmatch((s or "").strip()))

def _parse_stop_open_target(raw: str) -> Tuple[str, str]:
    s = (raw or "").strip()
    low = s.lower()
    for pref in ("username ", "user ", "handle "):
        if low.startswith(pref):
            t = s[len(pref):].strip()
            if not t.startswith("@"): t = f"@{t}"
            return ("username", t)
    for pref in ("whatsapp ", "wa ", "phone ", "number ", "num "):
        if low.startswith(pref):
            return ("phone", s[len(pref):].strip())
    if s.startswith("@"):
        return ("username", s)
    if _looks_like_phone(s):
        return ("phone", s)`
    return ("owner", s)

# =============================
# AUDIT LOG (DB)
# =============================
async def _log_event(kind: str, action: str, update: Update, value: str, owner: str = ""):
    try:
        u = update.effective_user
        m = update.effective_message
        c = update.effective_chat
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO audit_log (
                    ts_local, chat_id, message_id, user_id, user_first,
                    user_username, kind, action, value, owner
                ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
            """,
                datetime.now(TIMEZONE), c.id if c else None, m.message_id if m else None,
                u.id if u else None, u.first_name if u else None, u.username if u else None,
                kind, action, value, owner or ""
            )
    except Exception as e:
        log.warning("Log write to DB failed: %s", e)

# =============================
# UTIL
# =============================
def cache_user_info(user):
    state.setdefault("user_names", {})[str(user.id)] = {
        "first_name": user.first_name or "",
        "username": user.username or ""
    }

def mention_user_html(user_id: int) -> str:
    info = state.setdefault("user_names", {}).get(str(user_id), {})
    name = info.get("first_name") or info.get("username") or str(user_id)
    return f'<a href="tg://user?id={user_id}">{name}</a>'

def _issued_bucket(kind: str) -> Dict[str, list]:
    return state.setdefault("issued", {}).setdefault(kind, {})

async def _set_issued(user_id: int, chat_id: int, kind: str, value: str, context_data: Optional[Dict] = None):
    bucket = _issued_bucket(kind)
    user_id_str = str(user_id)
    if user_id_str not in bucket:
        bucket[user_id_str] = []

    item_data = {
        "value": value,
        "ts": datetime.now(TIMEZONE).isoformat(),
        "chat_id": chat_id
    }
    if context_data:
        item_data.update(context_data)

    bucket[user_id_str].append(item_data)
    await save_state()

async def _clear_issued(user_id: int, kind: str, value_to_clear: str) -> bool:
    bucket = _issued_bucket(kind)
    user_id_str = str(user_id)
    if user_id_str in bucket:
        original_len = len(bucket[user_id_str])
        bucket[user_id_str] = [
            item for item in bucket[user_id_str] if item.get("value") != value_to_clear
        ]
        if not bucket[user_id_str]:
            del bucket[user_id_str]

        if len(bucket.get(user_id_str, [])) < original_len:
            await save_state()
            return True
    return False

async def _clear_one_issued(user_id: int, kind: str, value_to_clear: str) -> bool:
    bucket = _issued_bucket(kind)
    user_id_str = str(user_id)
    if user_id_str in bucket:
        item_to_remove = None
        for item in bucket[user_id_str]:
            if item.get("value") == value_to_clear:
                item_to_remove = item
                break

        if item_to_remove:
            bucket[user_id_str].remove(item_to_remove)
            if not bucket[user_id_str]:
                del bucket[user_id_str]
            await save_state()
            return True
    return False

def _value_in_text(value: Optional[str], text: str) -> bool:
    if not value:
        return False

    v_norm = value.strip()
    text_norm = text or ""

    if v_norm.startswith('@'):
        pattern = re.compile(r'(?<!\S)' + re.escape(v_norm) + r'(?!\S)')
        return bool(pattern.search(text_norm))
    else:
        v_digits = re.sub(r'\D', '', v_norm)
        text_digits = re.sub(r'\D', '', text_norm)
        return v_digits and v_digits in text_digits

def _find_closest_app_id(typed_id: str) -> Optional[str]:
    """Finds the most similar pending App ID using Levenshtein distance."""

    def levenshtein(s1, s2):
        if len(s1) < len(s2):
            return levenshtein(s2, s1)
        if len(s2) == 0:
            return len(s1)
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        return previous_row[-1]

    all_pending_ids = []
    for _, items in _issued_bucket("app_id").items():
        for item in items:
            if item.get("value"):
                all_pending_ids.append(item.get("value"))

    if not all_pending_ids:
        return None

    norm_typed_id = _normalize_app_id(typed_id)

    closest_id = None
    min_distance = 3 # Max typo distance

    for pending_id in all_pending_ids:
        norm_pending_id = _normalize_app_id(pending_id)
        distance = levenshtein(norm_typed_id, norm_pending_id)

        if distance < min_distance:
            min_distance = distance
            closest_id = pending_id

    return closest_id if min_distance < 3 else None

# =============================
# COUNTRY & AGE FILTERING
# =============================
def _find_age_in_text(text: str) -> Optional[int]:
    match = re.search(
        r'\b(?:age|old)\s*:?\s*(\d{1,2})\b|\b(\d{1,2})\s*(?:yrs|yr|years|year old)\b',
        text.lower()
    )
    if match:
        age_str = match.group(1) or match.group(2)
        if age_str:
            return int(age_str)
    return None

def _find_country_in_text(text: str) -> Tuple[Optional[str], Optional[str]]:
    # MODIFIED: Reworked logic to be more robust
    match = re.search(r'\b(?:from|country)\s*:?\s*(.*)', text, re.IGNORECASE)
    if not match:
        return None, None

    line_after_from = match.group(1).split('\n')[0].strip().lower()

    for country in ALLOWED_COUNTRIES:
        if re.search(r'\b' + re.escape(country) + r'\b', line_after_from):
            if country in ['indian', 'india']:
                return country, 'india'
            return country, country

    potential_country_guess = line_after_from.split(',')[0].strip()
    return potential_country_guess, 'not_allowed'

# =============================
# DETAIL & PERFORMANCE COMMANDS
# =============================
async def _get_user_country_counts(user_id: int) -> List[Tuple[str, int]]:
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch(
                "SELECT country, count FROM user_daily_country_counts WHERE day=$1 AND user_id=$2 ORDER BY country",
                _logical_day_today(), user_id
            )
            return [(row['country'], row['count']) for row in rows]
    except Exception as e:
        log.warning(f"User country count read failed for {user_id}: {e}")
        return []

async def _get_user_confirmation_count(user_id: int) -> int:
    """Fetches a user's successful confirmation count for the current logical day."""
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            count = await conn.fetchval(
                "SELECT confirm_count FROM user_daily_confirmations WHERE day=$1 AND user_id=$2",
                _logical_day_today(), user_id
            )
            return int(count) if count is not None else 0
    except Exception as e:
        log.warning(f"User confirmation count read failed for {user_id}: {e}")
        return 0

async def _get_owner_performance(owner_name: str, day: date) -> Tuple[int, int]:
    """Fetches an owner's performance stats for a given day."""
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT telegram_count, whatsapp_count FROM owner_daily_performance WHERE day=$1 AND owner_name=$2",
                day, owner_name
            )
            return (row['telegram_count'], row['whatsapp_count']) if row else (0, 0)
    except Exception as e:
        log.warning(f"Owner performance read failed for {owner_name}: {e}")
        return (0, 0)

async def _get_owner_distribution_counts(owner_name: str, day: date) -> Tuple[int, int]:
    """Fetches an owner's distribution stats for a given day from the audit log."""
    start_ts = TIMEZONE.localize(datetime.combine(day, time(5, 30)))
    end_ts = start_ts + timedelta(days=1)
    tg_count = 0
    wa_count = 0
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch(
                """
                SELECT kind, COUNT(*) as count
                FROM audit_log
                WHERE owner = $1
                  AND action = 'issued'
                  AND ts_local >= $2 AND ts_local < $3
                GROUP BY kind;
                """,
                owner_name, start_ts, end_ts
            )
            for row in rows:
                if row['kind'] == 'username':
                    tg_count = row['count']
                elif row['kind'] == 'whatsapp':
                    wa_count = row['count']
            return (tg_count, wa_count)
    except Exception as e:
        log.warning(f"Owner distribution count read failed for {owner_name}: {e}")
        return (0, 0)

async def _get_user_performance_text(day: date) -> str:
    """Generates a ranked list of users by their successful customer confirmations."""
    lines = [f"<b>🏆 User Performance for {day.isoformat()}</b>"]

    user_performances = []

    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch(
                """
                SELECT user_id, confirm_count
                FROM user_daily_confirmations
                WHERE day = $1 AND confirm_count > 0
                ORDER BY confirm_count DESC;
                """,
                day
            )

            for row in rows:
                user_id = row['user_id']
                count = row['confirm_count']

                user_info = state.get("user_names", {}).get(str(user_id), {})
                user_display = user_info.get('username') or user_info.get('first_name') or f"ID: {user_id}"
                if user_info.get('username'):
                    user_display = f"@{user_display}"

                user_performances.append({'name': user_display, 'count': count})

    except Exception as e:
        log.error(f"Failed to get user performance data: {e}")
        return "An error occurred while fetching user performance data."

    if not user_performances:
        return f"No users added any customers on {day.isoformat()}."

    rank = 1
    for perf in user_performances:
        lines.append(f"<b>{rank}.</b> {perf['name']}: {perf['count']} customers")
        rank += 1

    return "\n".join(lines)

async def _get_user_stats_text(day: date) -> str:
    """Generates a ranked list of users by their success rate."""
    lines = [f"<b>📊 User Success Rate for {day.isoformat()}</b>"]
    user_stats = []

    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            # Fetch all activity and confirmations for the day
            activity_rows = await conn.fetch("SELECT user_id, username_requests, whatsapp_requests FROM user_daily_activity WHERE day = $1", day)
            confirm_rows = await conn.fetch("SELECT user_id, confirm_count FROM user_daily_confirmations WHERE day = $1", day)

            # Process into dictionaries for easy lookup
            activities = {r['user_id']: {'u': r['username_requests'], 'w': r['whatsapp_requests']} for r in activity_rows}
            confirmations = {r['user_id']: r['confirm_count'] for r in confirm_rows}

            # Get a set of all unique user IDs who were active
            all_user_ids = set(activities.keys()) | set(confirmations.keys())

            if not all_user_ids:
                return f"No user activity found for {day.isoformat()}."

            for user_id in all_user_ids:
                activity = activities.get(user_id, {'u': 0, 'w': 0})
                confirm_count = confirmations.get(user_id, 0)

                total_reqs = activity['u'] + activity['w']

                if total_reqs > 0:
                    rate = (confirm_count / total_reqs) * 100
                else:
                    rate = 0.0 # User had confirmations but no recorded requests (unlikely but possible)

                user_info = state.get("user_names", {}).get(str(user_id), {})
                user_display = user_info.get('username') or user_info.get('first_name') or f"ID: {user_id}"
                if user_info.get('username'):
                    user_display = f"@{user_display}"

                user_stats.append({
                    'name': user_display,
                    'rate': rate,
                    'confirm': confirm_count,
                    'reqs': total_reqs
                })

    except Exception as e:
        log.error(f"Failed to get user stats data: {e}")
        return "An error occurred while fetching user stats."

    # Sort users by success rate, from highest to lowest
    user_stats.sort(key=lambda x: x['rate'], reverse=True)

    rank = 1
    for stats in user_stats:
        lines.append(f"<b>{rank}.</b> {stats['name']}: <b>{stats['rate']:.1f}%</b> ({stats['confirm']} / {stats['reqs']})")
        rank += 1

    return "\n".join(lines)


async def _get_user_detail_text(user_id: int) -> str:
    user_info = state.get("user_names", {}).get(str(user_id), {})
    user_display = user_info.get('username') or user_info.get('first_name') or f"ID: {user_id}"
    if user_info.get('username'):
        user_display = f"@{user_display}"

    username_reqs, whatsapp_reqs = await _get_user_activity(user_id)
    country_counts = await _get_user_country_counts(user_id)
    confirmation_count = await _get_user_confirmation_count(user_id)

    pending_usernames = [item['value'] for item in _issued_bucket("username").get(str(user_id), [])]
    pending_whatsapps = [item['value'] for item in _issued_bucket("whatsapp").get(str(user_id), [])]

    lines = [f"<b>📊 Daily Detail for {user_display}</b>"]
    lines.append(f"<b>- Usernames Received:</b> {username_reqs}")
    lines.append(f"<b>- WhatsApps Received:</b> {whatsapp_reqs}")
    lines.append(f"<b>- Customers Added:</b> {confirmation_count}")

    if country_counts:
        lines.append("")
        lines.append("<b>🌍 Country Submissions:</b>")
        for country, count in country_counts:
            lines.append(f"  - {country.title()}: {count}")

    lines.append("")

    if pending_usernames:
        lines.append("<b>⏳ Pending Usernames:</b>")
        for u in pending_usernames:
            lines.append(f"  - <code>{u}</code>")
    else:
        lines.append("<b>✅ No Pending Usernames</b>")

    if pending_whatsapps:
        lines.append("\n<b>⏳ Pending WhatsApps:</b>")
        for w in pending_whatsapps:
            lines.append(f"  - <code>{w}</code>")
    else:
        lines.append("\n<b>✅ No Pending WhatsApps</b>")

    return "\n".join(lines)


async def _get_owner_performance_text(owner_name: str, day: date) -> str:
    """Generates a formatted string of an owner's daily performance and inventory."""
    # Daily Performance (from DB)
    tg_confirm_count, wa_confirm_count = await _get_owner_performance(owner_name, day)
    total_customers = tg_confirm_count + wa_confirm_count

    # NEW: Get distribution counts
    tg_dist_count, wa_dist_count = await _get_owner_distribution_counts(owner_name, day)

    lines = [f"<b>📊 Performance for @{owner_name} on {day.isoformat()}</b>"]
    lines.append(f"<b>- Customers via Telegram:</b> {tg_confirm_count}")
    lines.append(f"<b>- Customers via WhatsApp:</b> {wa_confirm_count}")
    lines.append(f"<b>- Total Customers Added:</b> {total_customers}")
    lines.append("") # Spacer
    lines.append("<b>🤖 Bot Distribution Stats</b>")
    lines.append(f"<b>- Usernames Sent from Bot:</b> {tg_dist_count}")
    lines.append(f"<b>- WhatsApps Sent from Bot:</b> {wa_dist_count}")
    lines.append("") # Spacer

    # Inventory Stats (from OWNER_DATA)
    owner_group = _find_owner_group(owner_name)
    total_tg = 0
    total_wa = 0
    stopped_tg = []
    stopped_wa = []

    if owner_group:
        total_tg = len(owner_group.get("entries", []))
        total_wa = len(owner_group.get("whatsapp", []))
        stopped_tg = [e.get("telegram") for e in owner_group.get("entries", []) if e.get("disabled")]
        stopped_wa = [w.get("number") for w in owner_group.get("whatsapp", []) if w.get("disabled")]

    lines.append("<b>📋 Current Inventory</b>")
    lines.append(f"<b>- Total Telegram in Bot:</b> {total_tg}")
    lines.append(f"<b>- Total WhatsApp in Bot:</b> {total_wa}")

    if stopped_tg:
        lines.append("")
        lines.append("<b>⛔ Stopped Telegram Usernames:</b>")
        for u in stopped_tg:
            lines.append(f"  - <code>{u}</code>")

    if stopped_wa:
        lines.append("")
        lines.append("<b>⛔ Stopped WhatsApp Numbers:</b>")
        for w in stopped_wa:
            lines.append(f"  - <code>{w}</code>")

    return "\n".join(lines)

# NEW: Function to generate the 'data today' report
async def _get_daily_data_summary_text() -> str:
    """Generates a summary of all owners who have added customers today."""
    today = _logical_day_today()
    lines = [f"<b>📊 Daily Customer Summary for {today.isoformat()}</b>"]

    owner_performances = []

    # Get all owner names from the loaded data
    all_owners = [_norm_owner_name(o['owner']) for o in OWNER_DATA]

    for owner_name in all_owners:
        tg_confirm_count, wa_confirm_count = await _get_owner_performance(owner_name, today)
        total_customers = tg_confirm_count + wa_confirm_count

        if total_customers > 0:
            owner_performances.append({'name': owner_name, 'total': total_customers})

    if not owner_performances:
        return "No owners have added customers today."

    # Sort by total customers, descending
    owner_performances.sort(key=lambda x: x['total'], reverse=True)

    for perf in owner_performances:
        lines.append(f"- @{perf['name']}: {perf['total']} customers")

    return "\n".join(lines)


# =============================
# EXCEL (reads audit_log)
# =============================
def _logical_day_of(ts: datetime) -> date:
    shifted = ts.astimezone(TIMEZONE) - timedelta(hours=5, minutes=30)
    return shifted.date()

async def _read_log_rows() -> List[dict]:
    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM audit_log ORDER BY ts_local")
            return [dict(row) for row in rows]
    except Exception as e:
        log.error("Failed to read log rows from DB: %s", e)
        return []

async def _compute_daily_summary(target_day: date) -> Tuple[List[dict], List[dict]]:
    """
    Computes daily summary for users and owners, now with added customer and country data.
    """
    # Define the time range for the logical day
    start_ts = TIMEZONE.localize(datetime.combine(target_day, time(5, 30)))
    end_ts = start_ts + timedelta(days=1)

    pool = await get_db_pool()
    async with pool.acquire() as conn:
        # Get all unique user IDs active on the target day from all relevant tables
        audit_users = await conn.fetch("SELECT DISTINCT user_id FROM audit_log WHERE ts_local >= $1 AND ts_local < $2 AND user_id IS NOT NULL", start_ts, end_ts)
        confirm_users = await conn.fetch("SELECT DISTINCT user_id FROM user_daily_confirmations WHERE day = $1", target_day)
        country_users = await conn.fetch("SELECT DISTINCT user_id FROM user_daily_country_counts WHERE day = $1", target_day)

        all_user_ids = {r['user_id'] for r in audit_users if r['user_id']}
        all_user_ids.update({r['user_id'] for r in confirm_users if r['user_id']})
        all_user_ids.update({r['user_id'] for r in country_users if r['user_id']})

        # Process data for each user
        out_users = []
        for user_id in all_user_ids:
            user_info = state.get("user_names", {}).get(str(user_id), {})
            user_display = user_info.get('username') or user_info.get('first_name') or f"ID: {user_id}"

            # Get request counts from audit log
            username_reqs, whatsapp_reqs = 0, 0
            req_rows = await conn.fetch("SELECT kind FROM audit_log WHERE ts_local >= $1 AND ts_local < $2 AND user_id = $3 AND action = 'issued'", start_ts, end_ts, user_id)
            for row in req_rows:
                if row['kind'] == 'username':
                    username_reqs += 1
                elif row['kind'] == 'whatsapp':
                    whatsapp_reqs += 1

            # Get confirmation counts
            confirm_count = await _get_user_confirmation_count(user_id)

            # Get country submissions
            country_counts = await _get_user_country_counts(user_id)
            country_str = ", ".join([f"{c.title()}: {n}" for c, n in country_counts])

            out_users.append({
                "Day": target_day.isoformat(),
                "User": user_display,
                "Total username receive": username_reqs,
                "Total whatsapp receive": whatsapp_reqs,
                "Total customer added": confirm_count,
                "Country Submissions": country_str,
            })

    # Owner performance logic remains largely the same, but re-fetched for clarity
    day_rows = [r for r in await _read_log_rows() if _logical_day_of(r["ts_local"]) == target_day]
    owner_stats: Dict[str, dict] = {}
    for r in day_rows:
        owner = (r.get("owner","") or "").lower()
        if r["action"] == "issued" and owner:
            s = owner_stats.setdefault(owner, {"total": 0, "tg": 0, "wa": 0})
            s["total"] += 1
            if r["kind"] == "username":
                s["tg"] += 1
            elif r["kind"] == "whatsapp":
                s["wa"] += 1

    out_owners = []
    for owner, s in sorted(owner_stats.items(), key=lambda kv: kv[0]):
        out_owners.append({
            "Day": target_day.isoformat(),
            "Owner": f"@{owner}",
            "Customers total": s["total"],
            "Customers via Telegram": s["tg"],
            "Customers via WhatsApp": s["wa"],
        })

    out_users.sort(key=lambda r: r["User"].lower())
    return out_users, out_owners

def _style_and_save_excel(user_rows: List[dict], owner_rows: List[dict]) -> io.BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed. Please add it to requirements.txt")

    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active; ws.title = "Summary"
    headers = ["Day", "User", "Total username receive", "Total whatsapp receive", "Total customer added", "Country Submissions"]
    ws.append(headers)
    for r in user_rows: ws.append([r.get(h, "") for h in headers])

    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = border
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
        fill = PatternFill(start_color="F2F2F2" if i % 2 == 0 else "FFFFFF", fill_type="solid")
        for cell in row: cell.alignment = center; cell.border = border; cell.fill = fill
    for col in ws.columns:
        max_len = 0; letter = col[0].column_letter
        for c in col:
            if c.value: max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)

    # Sheet 2: Owners
    ws2 = wb.create_sheet("Owners")
    headers2 = ["Day","Owner","Customers total","Customers via Telegram","Customers via WhatsApp"]
    ws2.append(headers2)
    for r in owner_rows: ws2.append([r.get(h, "") for h in headers2])
    for cell in ws2[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = border
    for i, row in enumerate(ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=ws2.max_column), start=2):
        fill = PatternFill(start_color="F2F2F2" if i % 2 == 0 else "FFFFFF", fill_type="solid")
        for cell in row: cell.alignment = center; cell.border = border; cell.fill = fill
    for col in ws2.columns:
        max_len = 0; letter = col[0].column_letter
        for c in col:
            if c.value: max_len = max(max_len, len(str(c.value)))
        ws2.column_dimensions[letter].width = min(max_len + 2, 40)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

async def _get_daily_excel_report(target_day: date) -> Tuple[Optional[str], Optional[io.BytesIO]]:
    user_rows, owner_rows = await _compute_daily_summary(target_day)
    if not user_rows and not owner_rows:
        return ("No data for that day.", None)
    try:
        excel_buffer = _style_and_save_excel(user_rows, owner_rows)
        return (None, excel_buffer)
    except Exception as e:
        log.error("Failed to generate Excel report in memory: %s", e)
        return (f"Failed to generate report: {e}", None)


def _parse_report_day(arg: Optional[str]) -> date:
    now = datetime.now(TIMEZONE)
    if not arg or arg.lower() == "today": return (now - timedelta(hours=5, minutes=30)).date()
    if arg.lower() == "yesterday": return (now - timedelta(hours=5, minutes=30, days=1)).date()
    try: return datetime.strptime(arg, "%Y-%m-%d").date()
    except Exception: return (now - timedelta(hours=5, minutes=30)).date()

# =============================
# ADMIN COMMANDS
# =============================
def _parse_duration(duration_str: str) -> Optional[timedelta]:
    m = re.match(r"(\d+)\s*(m|h|d|w)", (duration_str or "").strip().lower())
    if not m: return None
    val, unit = m.groups(); val = int(val)
    return {"m": timedelta(minutes=val), "h": timedelta(hours=val),
            "d": timedelta(days=val), "w": timedelta(weeks=val)}[unit]

def _find_owner_group(name: str) -> Optional[dict]:
    norm_name = _norm_owner_name(name)
    for group in OWNER_DATA:
        if _norm_owner_name(group["owner"]) == norm_name: return group
    return None

def _find_user_id_by_name(name: str) -> Optional[int]:
    norm_name = name.lower().lstrip('@').strip()
    user_names = state.get("user_names", {})
    for uid, data in user_names.items():
        if data.get("username", "").lower() == norm_name:
            return int(uid)
        if data.get("first_name", "").lower() == norm_name:
            return int(uid)
    return None

def _get_commands_text() -> str:
    return """
<b>Bot Command List</b>

<b>--- User Commands ---</b>
<code>i need username</code> - Request a username.
<code>i need whatsapp</code> - Request a WhatsApp number.
<code>my detail</code> - See your own daily stats.
<code>who's using @item</code> - Check the owner of an item.
<code>my performance</code> - See your daily customer stats (owners only).

<b>--- Admin: Owner & Item Management ---</b>
<code>add owner @owner</code>
<code>delete owner @owner</code>
<code>add username @user to @owner</code>
<code>delete username @user</code>
<code>add whatsapp +123... to @owner</code>
<code>delete whatsapp +123...</code>

<b>--- Admin: Availability Control ---</b>
<code>stop @owner/@user/+123...</code>
<code>open @owner/@user/+123...</code>
<code>stop all usernames</code>
<code>open all usernames</code>
<code>stop all whatsapp</code>
<code>open all whatsapp</code>
<code>stop all owners</code>
<code>open all owners</code>

<b>--- Admin: Priority & User Management ---</b>
<code>take 5 customer to owner @owner</code>
<code>take 5 customer to owner @owner and stop</code>
<code>ban whatsapp @user</code>
<code>unban whatsapp @user</code>
<code>list banned</code>
<code>ban country [country] for @user</code>
<code>unban country [country] for @user</code>
<code>list country bans [@user]</code>

<b>--- Admin: Reports & Manual Actions ---</b>
<code>report [today|yesterday|YYYY-MM-DD]</code>
<code>owner report [today|yesterday|YYYY-MM-DD]</code>
<code>user performance [day]</code> - See ranked list of users by customers added.
<code>user stats [day]</code> - See user success rates.
<code>request stats</code> - See current request ratio for auto-shutdown.
<code>performance @owner [day]</code> - See owner's customer stats.
<code>remind user</code>
<code>clear pending @item_or_number</code>
<code>data today</code> - Show today's customer summary by owner.

<b>--- Admin: Viewing Information ---</b>
<code>list owners</code>
<code>list disabled</code>
<code>list enabled</code> - List all active owners.
<code>list @owner</code>
<code>detail @user</code> - See a user's daily stats.
<code>inventory</code> - See total counts of active/paused items.

<b>--- Super Admin ---</b>
<code>add admin @user</code>
<code>delete admin @user</code>
<code>allow @user to use command [command]</code>
<code>stop allow @user to use command [command]</code>
<code>list admins</code>
<code>add user @user</code>
<code>delete user @user</code>
"""

async def _get_request_stats_text() -> str:
    """Generates a real-time report on the request ratio for the current 60-minute block."""
    now = datetime.now(TIMEZONE)
    # Use fixed 60-minute blocks (always starts at the top of the hour)
    start_time = now.replace(minute=0, second=0, microsecond=0)

    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            wa_count = await conn.fetchval(
                "SELECT COUNT(*) FROM audit_log WHERE kind = 'whatsapp' AND action = 'issued' AND ts_local >= $1",
                start_time
            )
            username_count = await conn.fetchval(
                "SELECT COUNT(*) FROM audit_log WHERE kind = 'username' AND action = 'issued' AND ts_local >= $1",
                start_time
            )
    except Exception as e:
        log.error(f"Failed to query audit log for request stats: {e}")
        return "Error fetching request stats from the database."

    minutes_until_next_block = 60 - now.minute
    block_start_str = start_time.strftime('%I:%M %p')

    lines = [f"<b>⏱️ Request Ratio Status (Block starts {block_start_str})</b>"]
    lines.append(f"<b>- Time Until Next Block:</b> {minutes_until_next_block} minutes")
    lines.append("")
    lines.append(f"<b>- Username Requests:</b> {username_count}")
    lines.append(f"<b>- WhatsApp Requests:</b> {wa_count}")
    lines.append("")

    if wa_count > username_count:
        status = "🔴 HIGH - WhatsApp numbers will be automatically stopped if this ratio persists."
    else:
        status = "🟢 NORMAL - Ratio is stable."

    lines.append(f"<b>- Current Status:</b> {status}")

    return "\n".join(lines)


def _get_inventory_text() -> str:
    """Generates a summary of the total number of items in the bot."""
    total_usernames = 0
    active_usernames = 0
    total_wa = 0
    active_wa = 0

    for owner_group in OWNER_DATA:
        for entry in owner_group.get("entries", []):
            total_usernames += 1
            if not entry.get("disabled"):
                active_usernames += 1
        for wa_entry in owner_group.get("whatsapp", []):
            total_wa += 1
            if not wa_entry.get("disabled"):
                active_wa += 1

    lines = ["<b>📋 Bot Inventory Summary</b>"]
    lines.append(f"<b>- Usernames:</b> {active_usernames} active / {total_usernames} total")
    lines.append(f"<b>- WhatsApp Numbers:</b> {active_wa} active / {total_wa} total")

    return "\n".join(lines)


async def _handle_admin_command(text: str, context: ContextTypes.DEFAULT_TYPE, update: Update) -> Optional[str]:
    user = update.effective_user

    # Super Admin Commands First
    if _is_super_admin(user):
        m_add_admin = ADD_ADMIN_RX.match(text)
        if m_add_admin:
            name = _norm_owner_name(m_add_admin.group(1))
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                await conn.execute("INSERT INTO admins (username, permissions) VALUES ($1, '[]') ON CONFLICT(username) DO NOTHING", name)
            await load_admins()
            return f"Admin '{name}' added with no permissions."

        m_add_user = ADD_USER_RX.match(text)
        if m_add_user:
            name = m_add_user.group(1)
            user_id = _find_user_id_by_name(name)
            if not user_id:
                return f"User '{name}' not found. They must interact with the bot once to be added."
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                await conn.execute("INSERT INTO whitelisted_users (user_id) VALUES ($1) ON CONFLICT (user_id) DO NOTHING", user_id)
            await load_whitelisted_users()
            return f"User '{name}' has been whitelisted."

        m_del_user = DELETE_USER_RX.match(text)
        if m_del_user:
            name = m_del_user.group(1)
            user_id = _find_user_id_by_name(name)
            if not user_id:
                return f"User '{name}' not found."
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                await conn.execute("DELETE FROM whitelisted_users WHERE user_id = $1", user_id)
            await load_whitelisted_users()
            return f"User '{name}' has been removed from the whitelist."


        m_del_admin = DELETE_ADMIN_RX.match(text)
        if m_del_admin:
            name = _norm_owner_name(m_del_admin.group(1))
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                await conn.execute("DELETE FROM admins WHERE username = $1", name)
            await load_admins()
            return f"Admin '{name}' deleted."

        m_allow = ALLOW_ADMIN_CMD_RX.match(text)
        if m_allow:
            name, command = m_allow.groups()
            name = _norm_owner_name(name)
            command = command.lower().strip()
            if command not in COMMAND_PERMISSIONS:
                return f"Invalid command name. Available commands: {', '.join(sorted(COMMAND_PERMISSIONS))}"

            current_perms = set(ADMIN_PERMISSIONS.get(name, []))
            current_perms.add(command)

            pool = await get_db_pool()
            async with pool.acquire() as conn:
                await conn.execute("INSERT INTO admins (username, permissions) VALUES ($1, $2) ON CONFLICT(username) DO UPDATE SET permissions = $2", name, json.dumps(list(current_perms)))
            await load_admins()
            return f"Permission '{command}' granted to '{name}'."

        m_stop_allow = STOP_ALLOW_ADMIN_CMD_RX.match(text)
        if m_stop_allow:
            name, command = m_stop_allow.groups()
            name = _norm_owner_name(name)
            command = command.lower().strip()
            if command not in COMMAND_PERMISSIONS:
                return f"Invalid command name. Available commands: {', '.join(sorted(COMMAND_PERMISSIONS))}"

            current_perms = set(ADMIN_PERMISSIONS.get(name, []))
            current_perms.discard(command)

            pool = await get_db_pool()
            async with pool.acquire() as conn:
                await conn.execute("UPDATE admins SET permissions = $1 WHERE username = $2", json.dumps(list(current_perms)), name)
            await load_admins()
            return f"Permission '{command}' revoked from '{name}'."

        if LIST_ADMINS_RX.match(text):
            if not ADMIN_PERMISSIONS:
                return "No admins configured."
            lines = ["<b>Configured Admins:</b>"]
            for name, perms in ADMIN_PERMISSIONS.items():
                lines.append(f"- <code>{name}</code>: {', '.join(perms) or 'No permissions'}")
            return "\n".join(lines)

    # Regular Admin Commands (with permission checks)
    m = TAKE_CUSTOMER_RX.match(text)
    if m:
        if not _has_permission(user, 'take customer'): return "You don't have permission to use this command."
        count_str, owner_name, and_stop_str = m.groups()
        count = int(count_str)
        owner_norm = _norm_owner_name(owner_name)

        owner_group = _find_owner_group(owner_norm)
        if not owner_group: return f"Owner '{owner_name}' not found."
        if _owner_is_paused(owner_group): return f"Owner '{owner_name}' is currently paused and cannot take customers."

        state["priority_queue"] = {
            "active": True, "owner": owner_norm, "remaining": count,
            "stop_after": bool(and_stop_str),
            "saved_rr_indices": { "username_owner_idx": state["rr"]["username_owner_idx"], "wa_owner_idx": state["rr"]["wa_owner_idx"] }
        }
        await save_state()
        stop_msg = " and will be stopped" if state["priority_queue"]["stop_after"] else ""
        return f"Priority queue activated: Next {count} customers will be directed to {owner_name}{stop_msg}."

    m = STOP_OPEN_RX.match(text)
    if m:
        if not _has_permission(user, 'stop open'): return "You don't have permission to use this command."
        action, target_raw = m.groups(); is_stop = action.lower() == "stop"
        current_admin = _norm_owner_name(user.username)
        t = target_raw.lower()
        if t in ("all whatsapp", "all whatsapps", "whatsapp all", "all wa", "wa all") or \
           t in ("all username", "all usernames", "username all", "usernames") or \
           t == "all owners":
            if not _is_super_admin(user):
                return "Only the super admin can perform 'stop/open all' actions."

        if t in ("all whatsapp", "all whatsapps", "whatsapp all", "all wa", "wa all"):
            total = changed = 0
            for owner in OWNER_DATA:
                for w in owner.get("whatsapp", []):
                    total += 1
                    if w.get("disabled") != is_stop: w["disabled"] = is_stop; changed += 1
            await _rebuild_pools_preserving_rotation()
            return f"{'Stopped' if is_stop else 'Opened'} all WhatsApp numbers — changed {changed}/{total}."

        if t in ("all username", "all usernames", "username all", "usernames"):
            total = changed = 0
            for owner in OWNER_DATA:
                for e in owner.get("entries", []):
                    total += 1
                    if e.get("disabled") != is_stop: e["disabled"] = is_stop; changed += 1
            await _rebuild_pools_preserving_rotation()
            return f"{'Stopped' if is_stop else 'Opened'} all usernames — changed {changed}/{total}."

        if t == "all owners":
            total = changed = 0
            for owner in OWNER_DATA:
                total += 1
                if owner.get("disabled", False) != is_stop:
                    owner["disabled"] = is_stop
                    owner.pop("disabled_until", None) # Also clear any timed pauses
                    changed += 1
            await _rebuild_pools_preserving_rotation()
            return f"{'Stopped' if is_stop else 'Opened'} all owners — changed {changed}/{total}."

        kind, value = _parse_stop_open_target(target_raw)
        if kind == "phone":
            norm_n = _norm_phone(value); found_item = None
            for owner in OWNER_DATA:
                for w in owner.get("whatsapp", []):
                    if _norm_phone(w.get("number")) == norm_n:
                        found_item = w
                        break
                if found_item: break

            if not found_item: return f"WhatsApp number {value} not found."

            manager = found_item.get("managed_by")
            if manager and manager != current_admin and not _is_super_admin(user):
                return f"You cannot manage this number. It is managed by @{manager}."

            found_item["disabled"] = is_stop
            if not _is_super_admin(user):
                found_item["managed_by"] = current_admin

            await _rebuild_pools_preserving_rotation()
            return f"{'Stopped' if is_stop else 'Opened'} WhatsApp {value}."
        if kind == "username":
            norm_h = _norm_handle(value); found_item = None
            for owner in OWNER_DATA:
                for e in owner.get("entries", []):
                    if _norm_handle(e.get("telegram")) == norm_h:
                        found_item = e
                        break
                if found_item: break

            if not found_item: return f"Username {value} not found."

            manager = found_item.get("managed_by")
            if manager and manager != current_admin and not _is_super_admin(user):
                return f"You cannot manage this username. It is managed by @{manager}."

            found_item["disabled"] = is_stop
            if not _is_super_admin(user):
                found_item["managed_by"] = current_admin

            await _rebuild_pools_preserving_rotation()
            return f"{'Stopped' if is_stop else 'Opened'} username {value}."
        owner = _find_owner_group(value)
        if not owner: return f"Owner '{value}' not found."

        manager = owner.get("managed_by")
        if manager and manager != current_admin and not _is_super_admin(user):
            return f"You cannot manage owner @{value}. It is managed by @{manager}."

        owner["disabled"] = bool(is_stop); owner.pop("disabled_until", None)
        if not _is_super_admin(user):
            owner["managed_by"] = current_admin
        await _rebuild_pools_preserving_rotation()
        return f"{'Stopped' if is_stop else 'Opened'} owner {value}."

    m = ADD_OWNER_RX.match(text)
    if m:
        if not _has_permission(user, 'add owner'): return "You don't have permission to use this command."
        name = _norm_owner_name(m.group(1))
        if _find_owner_group(name): return f"Owner '{name}' already exists."

        current_admin = _norm_owner_name(user.username)
        new_owner_data = {"owner": name}
        if not _is_super_admin(user):
            new_owner_data["managed_by"] = current_admin

        OWNER_DATA.append(_ensure_owner_shape(new_owner_data))
        await _rebuild_pools_preserving_rotation()
        return f"Owner '{name}' added."

    m = DEL_OWNER_RX.match(text)
    if m:
        if not _has_permission(user, 'delete owner'): return "You don't have permission to use this command."
        name = _norm_owner_name(m.group(1)); before = len(OWNER_DATA)
        owner_group_to_delete = _find_owner_group(name)
        if not owner_group_to_delete: return f"Owner '{name}' not found."

        current_admin = _norm_owner_name(user.username)
        manager = owner_group_to_delete.get("managed_by")
        if manager and manager != current_admin and not _is_super_admin(user):
            return f"You cannot delete owner @{name}. It is managed by @{manager}."

        OWNER_DATA[:] = [g for g in OWNER_DATA if _norm_owner_name(g.get("owner","")) != name]
        if len(OWNER_DATA) == before: return f"Owner '{name}' not found."
        await _rebuild_pools_preserving_rotation()
        return f"Owner '{name}' deleted."

    m = ADD_USERNAME_RX.match(text)
    if m:
        if not _has_permission(user, 'add username'): return "You don't have permission to use this command."
        handle, owner_name = m.groups(); owner = _find_owner_group(owner_name)
        if not owner: return f"Owner '{owner_name}' not found."
        norm_h = _norm_handle(handle)
        if any(_norm_handle(e.get("telegram")) == norm_h for e in owner["entries"]): return f"@{handle} already exists for owner {owner_name}."

        current_admin = _norm_owner_name(user.username)
        new_entry = {"telegram": handle, "phone": "", "disabled": False}
        if not _is_super_admin(user):
            new_entry["managed_by"] = current_admin

        owner["entries"].append(new_entry)
        await _rebuild_pools_preserving_rotation()
        return f"Added username @{handle} to {owner_name}."

    m = ADD_WHATSAPP_RX.match(text)
    if m:
        if not _has_permission(user, 'add whatsapp'): return "You don't have permission to use this command."
        num, owner_name = m.groups(); owner = _find_owner_group(owner_name)
        if not owner: return f"Owner '{owner_name}' not found."
        norm_n = _norm_phone(num)
        if any(_norm_phone(w.get("number")) == norm_n for w in owner["whatsapp"]): return f"Number {num} already exists for owner {owner_name}."

        current_admin = _norm_owner_name(user.username)
        new_wa = {"number": num, "disabled": False}
        if not _is_super_admin(user):
            new_wa["managed_by"] = current_admin

        owner["whatsapp"].append(new_wa)
        await _rebuild_pools_preserving_rotation()
        return f"Added WhatsApp {num} to {owner_name}."

    m = DEL_USERNAME_RX.match(text)
    if m:
        if not _has_permission(user, 'delete username'): return "You don't have permission to use this command."
        handle = m.group(1); norm_h = _norm_handle(handle); found_and_deleted = False
        current_admin = _norm_owner_name(user.username)
        for owner in OWNER_DATA:
            entry_to_delete = None
            for e in owner["entries"]:
                if _norm_handle(e.get("telegram")) == norm_h:
                    entry_to_delete = e
                    break

            if entry_to_delete:
                manager = entry_to_delete.get("managed_by")
                if manager and manager != current_admin and not _is_super_admin(user):
                    return f"You cannot delete username @{handle}. It is managed by @{manager}."

                owner["entries"].remove(entry_to_delete)
                found_and_deleted = True

        if not found_and_deleted: return f"Username @{handle} not found."
        await _rebuild_pools_preserving_rotation()
        return f"Deleted username @{handle} from all owners."

    m = DEL_WHATSAPP_RX.match(text)
    if m:
        if not _has_permission(user, 'delete whatsapp'): return "You don't have permission to use this command."
        num = m.group(1); norm_n = _norm_phone(num); found_and_deleted = False
        current_admin = _norm_owner_name(user.username)
        for owner in OWNER_DATA:
            wa_to_delete = None
            for w in owner["whatsapp"]:
                if _norm_phone(w.get("number")) == norm_n:
                    wa_to_delete = w
                    break

            if wa_to_delete:
                manager = wa_to_delete.get("managed_by")
                if manager and manager != current_admin and not _is_super_admin(user):
                    return f"You cannot delete number {num}. It is managed by @{manager}."

                owner["whatsapp"].remove(wa_to_delete)
                found_and_deleted = True

        if not found_and_deleted: return f"WhatsApp number {num} not found."
        await _rebuild_pools_preserving_rotation()
        return f"Deleted WhatsApp number {num} from all owners."

    if LIST_OWNERS_RX.match(text):
        if not _has_permission(user, 'list owners'): return "You're not authorized to use this command."
        if not OWNER_DATA: return "No owners configured."
        lines = ["<b>Owner Roster:</b>"]
        for o in OWNER_DATA:
            status = "PAUSED" if _owner_is_paused(o) else "active"
            u_count = len([e for e in o.get("entries", []) if not e.get("disabled")])
            w_count = len([w for w in o.get("whatsapp", []) if not w.get("disabled")])
            lines.append(f"- <code>{o['owner']}</code> ({status}): {u_count} usernames, {w_count} whatsapps")
        return "\n".join(lines)

    if LIST_DISABLED_RX.match(text):
        if not _has_permission(user, 'list disabled'): return "You're not authorized to use this command."
        disabled = [o for o in OWNER_DATA if _owner_is_paused(o)]
        if not disabled: return "No owners are currently disabled/paused."
        lines = ["<b>Disabled/Paused Owners:</b>"]
        for o in disabled:
            reason = f" (until {o['disabled_until']})" if o.get("disabled_until") else ""
            lines.append(f"- <code>{o['owner']}</code>{reason}")
        return "\n".join(lines)

    m = LIST_ENABLED_RX.match(text)
    if m:
        if not _has_permission(user, 'list enabled'): return "You're not authorized to use this command."
        enabled = [o for o in OWNER_DATA if not _owner_is_paused(o)]
        if not enabled: return "No owners are currently enabled/active."
        lines = ["<b>Enabled/Active Owners:</b>"]
        for o in enabled:
            u_count = len([e for e in o.get("entries", []) if not e.get("disabled")])
            w_count = len([w for w in o.get("whatsapp", []) if not w.get("disabled")])
            lines.append(f"- <code>{o['owner']}</code>: {u_count} usernames, {w_count} whatsapps")
        return "\n".join(lines)


    m = LIST_OWNER_DETAIL_RX.match(text) or LIST_OWNER_ALIAS_RX.match(text)
    if m:
        if not _has_permission(user, 'list owner'): return "You're not authorized to use this command."
        name = m.group(1).strip()
        if name.lower() in ("owners", "disabled"): return None
        owner = _find_owner_group(name)
        if not owner: return f"Owner '{name}' not found."

        owner_status_flag = " ⛔" if _owner_is_paused(owner) else ""
        lines = [f"<b>Details for {owner['owner']}{owner_status_flag}:</b>"]
        if owner.get("entries"):
            lines.append("<u>Usernames:</u>")
            for e in owner["entries"]:
                flag = " ⛔" if e.get("disabled") else ""; h = e.get("telegram") or ""
                if h and not h.startswith("@"): h = "@" + h
                lines.append(f"- {h}{flag}")
        if owner.get("whatsapp"):
            lines.append("<u>WhatsApp Numbers:</u>")
            for w in owner["whatsapp"]:
                flag = " ⛔" if w.get("disabled") else ""
                lines.append(f"- {w['number']}{flag}")
        if len(lines) == 1: lines.append("No entries found.")
        return "\n".join(lines)

    m = REMIND_ALL_RX.match(text)
    if m:
        if not _has_permission(user, 'remind user'): return "You don't have permission to use this command."
        return await _send_all_pending_reminders(context)

    m = CLEAR_PENDING_RX.match(text)
    if m:
        if not _has_permission(user, 'clear pending'): return "You don't have permission to use this command."
        if not _is_super_admin(user): return "Only the super admin can use this command."

        item_to_clear = m.group(1).strip()
        for kind in ("username", "whatsapp", "app_id"):
            for user_id_str, items in list(_issued_bucket(kind).items()):
                for item in items:
                    stored_value = item.get("value")
                    match_found = False
                    if kind in ("username", "app_id"):
                        if stored_value and item_to_clear.lower() == stored_value.lower():
                            match_found = True
                    elif kind == "whatsapp":
                        if stored_value and _norm_phone(item_to_clear) == _norm_phone(stored_value):
                            match_found = True

                    if match_found:
                        user_id = int(user_id_str)
                        if await _clear_issued(user_id, kind, stored_value):
                            user_info = state.get("user_names", {}).get(user_id_str, {})
                            user_name = user_info.get("username") or user_info.get("first_name") or f"ID {user_id}"
                            reply_message = f"✅ Cleared pending {kind} <code>{stored_value}</code> for user {user_name}."

                            # If a WA item is cleared by a super admin, lift any related temp ban
                            if kind == "whatsapp":
                                temp_bans = state.setdefault("whatsapp_temp_bans", {})
                                if user_id_str in temp_bans:
                                    del temp_bans[user_id_str]
                                    await save_state() # Save state after modifying bans
                                    reply_message += "\nUser's temporary WhatsApp ban has been lifted."
                                    log.info(f"Super admin cleared pending WA, lifting temp ban for user {user_id_str}.")

                            return reply_message

        return f"❌ Could not find any user with the pending item <code>{item_to_clear}</code>."

    m = BAN_WHATSAPP_RX.match(text)
    if m:
        if not _has_permission(user, 'ban whatsapp'): return "You don't have permission to use this command."
        target_name = m.group(1)
        user_id_to_ban = _find_user_id_by_name(target_name)
        if not user_id_to_ban:
            return f"User '{target_name}' not found."

        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("INSERT INTO whatsapp_bans (user_id) VALUES ($1) ON CONFLICT (user_id) DO NOTHING", user_id_to_ban)

        WHATSAPP_BANNED_USERS.add(user_id_to_ban)
        return f"User {target_name} has been banned from requesting WhatsApp numbers."

    m = UNBAN_WHATSAPP_RX.match(text)
    if m:
        if not _has_permission(user, 'unban whatsapp'): return "You don't have permission to use this command."
        target_name = m.group(1)
        user_id_to_unban = _find_user_id_by_name(target_name)
        # Check if user_id_to_unban exists before proceeding
        if not user_id_to_unban:
            return f"User '{target_name}' not found."

        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("DELETE FROM whatsapp_bans WHERE user_id = $1", user_id_to_unban)

        WHATSAPP_BANNED_USERS.discard(user_id_to_unban)
        return f"User {target_name} has been unbanned from requesting WhatsApp numbers."

    if LIST_BANNED_RX.match(text):
        if not _has_permission(user, 'list banned'): return "You're not authorized to use this command."
        if not WHATSAPP_BANNED_USERS:
            return "No users are currently banned from requesting WhatsApp numbers."

        lines = ["<b>WhatsApp Banned Users:</b>"]
        for user_id in WHATSAPP_BANNED_USERS:
            user_info = state.get("user_names", {}).get(str(user_id), {})
            user_display = user_info.get('username') or user_info.get('first_name') or f"ID: {user_id}"
            lines.append(f"- {user_display}")
        return "\n".join(lines)

    m = BAN_COUNTRY_RX.match(text)
    if m:
        if not _has_permission(user, 'ban country'): return "You're not authorized to use this command."
        country_raw, target_name = m.groups()
        country = country_raw.strip().lower()

        target_user_id = _find_user_id_by_name(target_name)
        if not target_user_id:
            return f"User '{target_name}' not found."

        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("INSERT INTO user_country_bans (user_id, country) VALUES ($1, $2) ON CONFLICT (user_id, country) DO NOTHING", target_user_id, country)

        await load_user_country_bans() # Reload the cache
        return f"User {target_name} is now MANUALLY banned from submitting for country: '{country}'."

    m = UNBAN_COUNTRY_RX.match(text)
    if m:
        if not _has_permission(user, 'unban country'): return "You're not authorized to use this command."
        country_raw, target_name = m.groups()
        country = country_raw.strip().lower()

        target_user_id = _find_user_id_by_name(target_name)
        if not target_user_id:
            return f"User '{target_name}' not found."

        pool = await get_db_pool()
        async with pool.acquire() as conn:
            await conn.execute("DELETE FROM user_country_bans WHERE user_id = $1 AND country = $2", target_user_id, country)

        await load_user_country_bans() # Reload the cache
        return f"User {target_name} has been MANUALLY unbanned from submitting for country: '{country}'."

    m = LIST_COUNTRY_BANS_RX.match(text)
    if m:
        if not _has_permission(user, 'list country bans'): return "You're not authorized to use this command."
        target_name = m.group(1)

        if target_name:
            target_user_id = _find_user_id_by_name(target_name)
            if not target_user_id:
                return f"User '{target_name}' not found."

            user_bans = USER_COUNTRY_BANS.get(target_user_id, set())
            if not user_bans:
                return f"User {target_name} has no manual country bans."

            lines = [f"<b>Manual country bans for {target_name}:</b>"]
            lines.extend(f"- {c.title()}" for c in sorted(list(user_bans)))
            return "\n".join(lines)
        else:
            if not USER_COUNTRY_BANS:
                return "No users have manual (permanent) country-specific bans."

            lines = ["<b>All User-Specific Manual Country Bans:</b>"]
            for user_id, banned_countries in sorted(USER_COUNTRY_BANS.items()):
                user_info = state.get("user_names", {}).get(str(user_id), {})
                user_display = user_info.get('username') or user_info.get('first_name') or f"ID: {user_id}"
                if user_info.get('username'):
                    user_display = f"@{user_display}"

                countries_str = ", ".join(sorted([c.title() for c in banned_countries]))
                lines.append(f"- <b>{user_display}</b>: {countries_str}")
            return "\n".join(lines)


    m = OWNER_REPORT_RX.match(text)
    if m:
        if not _has_permission(user, 'owner report'): return "You don't have permission to use this command."
        target_day = _parse_report_day(m.group(1))
        _, owner_rows = await _compute_daily_summary(target_day)
        if not owner_rows:
            return f"No owner activity found for {target_day.isoformat()}."

        lines = [f"<b>Owner Performance for {target_day.isoformat()}:</b>"]
        for row in owner_rows:
            lines.append(f"- <b>{row['Owner']}</b>: {row['Customers total']} ({row['Customers via Telegram']} usernames, {row['Customers via WhatsApp']} whatsapps)")
        return "\n".join(lines)

    m = USER_PERFORMANCE_RX.match(text)
    if m:
        if not _has_permission(user, 'user performance'): return "You're not authorized to use this command."
        target_day = _parse_report_day(m.group(1))
        return await _get_user_performance_text(target_day)

    m = USER_STATS_RX.match(text)
    if m:
        if not _has_permission(user, 'user stats'): return "You're not authorized to use this command."
        target_day = _parse_report_day(m.group(1))
        return await _get_user_stats_text(target_day)

    if INVENTORY_RX.match(text):
        if not _has_permission(user, 'inventory'): return "You're not authorized to use this command."
        return _get_inventory_text()

    if REQUEST_STATS_RX.match(text):
        if not _has_permission(user, 'request stats'): return "You're not authorized to use this command."
        return await _get_request_stats_text()

    if COMMANDS_RX.match(text):
        command_list_text = _get_commands_text()
        return command_list_text

    m = DATA_TODAY_RX.match(text)
    if m:
        if not _has_permission(user, 'data today'):
            return "You're not authorized to use this command."
        return await _get_daily_data_summary_text()

    m = DETAIL_USER_RX.match(text)
    if m:
        if not _has_permission(user, 'detail user'): return "You're not authorized to use this command."
        target_name = m.group(1)
        target_user_id = _find_user_id_by_name(target_name)
        if not target_user_id:
            return f"User '{target_name}' not found."
        return await _get_user_detail_text(target_user_id)


    return None

# =============================
# REMINDER & RESET TASKS
# =============================
async def _send_all_pending_reminders(context: ContextTypes.DEFAULT_TYPE) -> str:
    total_reminders_sent = 0
    reminded_users = set()

    reminders_to_send = []
    # MODIFIED: Removed 'app_id' from the reminder loop
    for kind in ("username", "whatsapp"):
        bucket = _issued_bucket(kind)
        for user_id_str, items in bucket.items():
            for item in items:
                user_id = int(user_id_str)
                chat_id = item.get("chat_id")
                value = item.get("value")
                if chat_id and value:
                    label = "username" if kind == "username" else "WhatsApp"
                    reminder_text = (
                        f"សូមរំលឹក: {mention_user_html(user_id)}, "
                        f"អ្នកនៅមិនទាន់បានផ្តល់ព័ត៌មានសម្រាប់ {label} {value} ដែលអ្នកបានស្នើសុំ។"
                    )
                    reminders_to_send.append({'chat_id': chat_id, 'text': reminder_text, 'user_id': user_id})

    if not reminders_to_send:
        return "No pending items found to send reminders for."

    for r in reminders_to_send:
        try:
            await context.bot.send_message(
                chat_id=r['chat_id'],
                text=r['text'],
                parse_mode=ParseMode.HTML
            )
            total_reminders_sent += 1
            reminded_users.add(r['user_id'])
        except Exception as e:
            log.error(f"Error sending manual reminder for user {r['user_id']}: {e}")

    return f"Successfully sent {total_reminders_sent} reminder(s) to {len(reminded_users)} user(s)."

# NEW: Function to clear expired App IDs
async def _clear_expired_app_ids(context: ContextTypes.DEFAULT_TYPE):
    log.info("Running hourly check for expired App IDs...")
    now = datetime.now(TIMEZONE)
    forty_eight_hours = timedelta(hours=48)
    state_changed = False

    pending_apps = _issued_bucket("app_id")

    for user_id_str, items in list(pending_apps.items()):
        items_to_keep = []
        for item in items:
            try:
                item_ts = datetime.fromisoformat(item["ts"])
                if (now - item_ts) > forty_eight_hours:
                    log.info(f"Expired App ID '{item['value']}' for user {user_id_str} removed after 48 hours.")
                    state_changed = True
                else:
                    items_to_keep.append(item)
            except Exception as e:
                log.warning(f"Could not parse timestamp for item {item} for user {user_id_str}: {e}")
                items_to_keep.append(item) # Keep item if timestamp is invalid

        if not items_to_keep:
            if user_id_str in pending_apps:
                del pending_apps[user_id_str]
        else:
            pending_apps[user_id_str] = items_to_keep

    if state_changed:
        await save_state()
        log.info("Finished clearing expired App IDs. State saved.")
    else:
        log.info("No expired App IDs found.")


async def check_reminders(context: ContextTypes.DEFAULT_TYPE):
    """
    Handles sending reminders and applying escalating bans for non-compliance.
    - Usernames get repeated reminders.
    - WhatsApp failures escalate from 30min ban -> 2hr ban -> permanent ban.
    - Overdue WhatsApp items are NOT cleared automatically, but are marked to prevent repeat punishments.
    """
    reminders_to_send = []
    state_changed = False

    async with db_lock:
        now = datetime.now(TIMEZONE)
        pool = await get_db_pool() # For permanent bans

        # Handle username reminders (sends a reminder every `REMINDER_DELAY_MINUTES`)
        username_bucket = _issued_bucket("username")
        for user_id_str, items in list(username_bucket.items()):
            for item in list(items):
                try:
                    last_reminder_ts_str = item.get("last_reminder_ts")
                    base_ts = datetime.fromisoformat(last_reminder_ts_str) if last_reminder_ts_str else datetime.fromisoformat(item["ts"])

                    if (now - base_ts) > timedelta(minutes=REMINDER_DELAY_MINUTES):
                        user_id = int(user_id_str)
                        chat_id = item.get("chat_id")
                        value = item.get("value")
                        if chat_id and value:
                            reminder_text = (
                                f"សូមរំលឹក: {mention_user_html(user_id)}, "
                                f"អ្នកនៅមិនទាន់បានផ្តល់ព័ត៌មានសម្រាប់ username {value} ដែលអ្នកបានស្នើសុំ។"
                            )
                            reminders_to_send.append({'chat_id': chat_id, 'text': reminder_text})
                            item["last_reminder_ts"] = now.isoformat()
                            state_changed = True
                            log.info(f"Queued recurring reminder for user {user_id} for username '{value}'")
                except Exception as e:
                    log.error(f"Error processing username reminder for user {user_id_str}: {e}")

        # Handle WhatsApp reminders and escalating bans
        whatsapp_bucket = _issued_bucket("whatsapp")
        for user_id_str, items in list(whatsapp_bucket.items()):
            user_id = int(user_id_str)
            for item in items:
                try:
                    item_ts = datetime.fromisoformat(item["ts"])

                    # Check if item is overdue AND has not already been punished for this offense
                    if (now - item_ts) > timedelta(minutes=REMINDER_DELAY_MINUTES) and not item.get("punished"):
                        # This item is overdue. Time to remind AND ban.
                        offense_counts = state.setdefault("whatsapp_offense_count", {})
                        offense_count = offense_counts.get(user_id_str, 0) + 1
                        offense_counts[user_id_str] = offense_count
                        log.info(f"User {user_id} failed to clear WA '{item.get('value')}'. Offense count is now {offense_count}.")

                        ban_message_khmer = ""
                        if offense_count == 1:
                            # First offense: 30-minute ban
                            ban_duration_minutes = 30
                            ban_until = now + timedelta(minutes=ban_duration_minutes)
                            state.setdefault("whatsapp_temp_bans", {})[user_id_str] = ban_until.isoformat()
                            ban_message_khmer = (
                                f"សូមរំលឹក: {mention_user_html(user_id)}, អ្នកមិនបានផ្តល់ព័ត៌មានសម្រាប់ WhatsApp {item.get('value')}\n"
                                f"អ្នកត្រូវបានហាមឃាត់ជាបណ្ដោះអាសន្នពីការស្នើសុំលេខ WhatsApp រយៈពេល {ban_duration_minutes} នាទី។"
                            )
                            log.info(f"User {user_id} temp-banned for {ban_duration_minutes} mins (1st offense).")
                        elif offense_count == 2:
                            # Second offense: 2-hour ban
                            ban_duration_minutes = 120
                            ban_until = now + timedelta(minutes=ban_duration_minutes)
                            state.setdefault("whatsapp_temp_bans", {})[user_id_str] = ban_until.isoformat()
                            ban_message_khmer = (
                                f"{mention_user_html(user_id)}, ដោយសារអ្នកបានធ្វើកំហុសដដែលម្តងទៀត, "
                                f"អ្នកត្រូវបានហាមឃាត់ពីការស្នើសុំលេខ WhatsApp រយៈពេល 2 ម៉ោង។"
                            )
                            log.info(f"User {user_id} temp-banned for {ban_duration_minutes} mins (2nd offense).")
                        else: # 3rd or more offense
                            # Third offense: Permanent ban
                            WHATSAPP_BANNED_USERS.add(user_id)
                            async with pool.acquire() as conn:
                                await conn.execute("INSERT INTO whatsapp_bans (user_id) VALUES ($1) ON CONFLICT (user_id) DO NOTHING", user_id)
                            ban_message_khmer = (
                                f"{mention_user_html(user_id)}, ដោយសារតែការមិនគោរពតាមការរំលឹកជាច្រើនដង, "
                                f"អ្នកត្រូវបានហាមឃាត់ជាអចិន្ត្រៃយ៍ពីការស្នើសុំលេខ WhatsApp។"
                            )
                            log.info(f"User {user_id} permanently banned from WhatsApp requests (3rd+ offense).")

                        if ban_message_khmer:
                            reminders_to_send.append({'chat_id': item.get("chat_id"), 'text': ban_message_khmer})

                        # Mark the item as punished to prevent re-punishing, but DO NOT clear it.
                        item["punished"] = True
                        state_changed = True

                except Exception as e:
                    log.error(f"Error processing reminder/ban for user {user_id_str}: {e}")

        if state_changed:
            await save_state()

    # Send all queued messages
    for r in reminders_to_send:
        try:
            await context.bot.send_message(chat_id=r['chat_id'], text=r['text'], parse_mode=ParseMode.HTML)
        except Exception as e:
            log.error(f"Failed to send reminder/ban message to chat {r['chat_id']}: {e}")

async def check_request_ratio_and_stop_whatsapp(context: ContextTypes.DEFAULT_TYPE):
    """Checks the ratio of WA to Username requests for the previous 60-min block and stops all WA if the ratio is too high."""
    log.info("Running 60-minute check of request ratio...")

    now = datetime.now(TIMEZONE)

    # Determine the start and end of the 60-minute block that just concluded.
    # The job runs at the top of the hour, so we check the previous hour.
    end_time = now.replace(minute=0, second=0, microsecond=0)
    start_time = end_time - timedelta(hours=1)

    try:
        pool = await get_db_pool()
        async with pool.acquire() as conn:
            # Query for the specific, fixed block
            wa_count = await conn.fetchval(
                "SELECT COUNT(*) FROM audit_log WHERE kind = 'whatsapp' AND action = 'issued' AND ts_local >= $1 AND ts_local < $2",
                start_time, end_time
            )
            username_count = await conn.fetchval(
                "SELECT COUNT(*) FROM audit_log WHERE kind = 'username' AND action = 'issued' AND ts_local >= $1 AND ts_local < $2",
                start_time, end_time
            )
    except Exception as e:
        log.error(f"Failed to query audit log for request ratio check: {e}")
        return

    log.info(f"Request ratio check for block {start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}: {wa_count} WhatsApps, {username_count} Usernames.")

    if wa_count > username_count:
        log.warning(f"WhatsApp requests ({wa_count}) exceeded username requests ({username_count}) in the last block. Stopping all WhatsApp numbers.")

        changed = 0
        async with db_lock:
            for owner in OWNER_DATA:
                for w_entry in owner.get("whatsapp", []):
                    if not w_entry.get("disabled"):
                        w_entry["disabled"] = True
                        changed += 1

            if changed > 0:
                await _rebuild_pools_preserving_rotation()

                notification_text = (
                    f"⚠️ <b>Automatic Action</b> ⚠️\n\n"
                    f"All WhatsApp numbers have been temporarily disabled because WhatsApp requests ({wa_count}) "
                    f"exceeded username requests ({username_count}) in the last 60-minute block.\n\n"
                    f"An admin can re-enable them using the <code>open all whatsapp</code> or <code>open [number]</code> command."
                )
                try:
                    await context.bot.send_message(chat_id=REQUEST_GROUP_ID, text=notification_text, parse_mode=ParseMode.HTML)
                except Exception as e:
                    log.error(f"Failed to send request ratio notification: {e}")


async def daily_reset(context: ContextTypes.DEFAULT_TYPE):
    log.info("Performing daily reset...")
    async with db_lock:
        state['whatsapp_offense_count'] = {} # Reset offense counts daily
        log.info("Resetting daily catch-up assignments, cooldowns, and WhatsApp offense counters.")
        try:
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                await conn.execute("DELETE FROM wa_daily_usage;")
                await conn.execute("DELETE FROM user_daily_activity;")
                await conn.execute("DELETE FROM user_daily_country_counts;")
                await conn.execute("DELETE FROM user_daily_confirmations;")
                await conn.execute("DELETE FROM owner_daily_performance;")
                # MODIFICATION: DO NOT clear user_country_bans, as this is for manual/permanent bans.
                # The automatic daily limit is enforced by clearing user_daily_country_counts.
                log.info("Cleared daily WhatsApp, user activity, country, confirmation, and performance quotas from database.")
        except Exception as e:
            log.error(f"Failed to clear daily tables: {e}")

        await save_state()
        log.info("Daily reset complete. Pending items from previous days are preserved.")

# =============================
# MESSAGE HANDLER
# =============================
async def on_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_chat or not update.effective_user or \
       update.effective_chat.type not in (ChatType.GROUP, ChatType.SUPERGROUP):
        return

    msg = update.effective_message
    text = (msg.text or msg.caption or "").strip()
    uid = update.effective_user.id
    chat_id = msg.chat_id
    cache_user_info(update.effective_user)

    async with db_lock:
        # User "my detail" command
        if MY_DETAIL_RX.match(text):
            if chat_id == DETAIL_GROUP_ID:
                if uid not in WHITELISTED_USERS and not _is_admin(update.effective_user):
                    return
                detail_text = await _get_user_detail_text(uid)
                await msg.reply_html(detail_text)
            return

        # Owner "my performance" command
        m_my_perf = MY_PERFORMANCE_RX.match(text)
        if m_my_perf:
            if chat_id in PERFORMANCE_GROUP_IDS:
                if _is_owner(update.effective_user):
                    owner_name = _norm_owner_name(update.effective_user.username)
                    target_day = _parse_report_day(m_my_perf.group(1))
                    perf_text = await _get_owner_performance_text(owner_name, target_day)
                    await msg.reply_html(perf_text)
                else:
                    await msg.reply_text("This command is only for registered owners.")
            return

        # Admin: Report
        mrep = SEND_REPORT_RX.match(text)
        if _is_admin(update.effective_user) and _has_permission(update.effective_user, 'report') and mrep:
            target_day = _parse_report_day(mrep.group(1))
            err, excel_buffer = await _get_daily_excel_report(target_day)

            if err:
                await msg.reply_text(err)
            elif excel_buffer:
                file_name = f"daily_summary_{target_day.isoformat()}.xlsx"
                await msg.reply_document(
                    document=excel_buffer, filename=file_name,
                    caption=f"Daily summary (logical day starting 05:30) — {target_day}"
                )
            return

        # Admin "performance @owner" command
        m_owner_perf = PERFORMANCE_OWNER_RX.match(text)
        if _is_admin(update.effective_user) and _has_permission(update.effective_user, 'performance') and m_owner_perf:
            owner_name_raw, day_str = m_owner_perf.groups()
            owner_name = _norm_owner_name(owner_name_raw)
            if not _find_owner_group(owner_name):
                await msg.reply_text(f"Owner '{owner_name}' not found.")
                return
            target_day = _parse_report_day(day_str)
            perf_text = await _get_owner_performance_text(owner_name, target_day)
            await msg.reply_html(perf_text)
            return

        # Admin: Console
        if _is_admin(update.effective_user):
            admin_reply = await _handle_admin_command(text, context, update)
            if admin_reply:
                await msg.reply_html(admin_reply)
                return

        if chat_id == CONFIRMATION_GROUP_ID:
            if '+1' in text:
                match = re.search(r'@([^\s]+)', text)
                if match:
                    app_id_confirmed_raw = f"@{match.group(1)}"
                    found_and_counted = False

                    # Search all users to find who this App ID belongs to
                    for user_id_str, items in list(_issued_bucket("app_id").items()):
                        if found_and_counted: break
                        for item in items:
                            stored_app_id_raw = item.get("value", "")

                            match_is_found = (_normalize_app_id(stored_app_id_raw) == _normalize_app_id(app_id_confirmed_raw))

                            if match_is_found:
                                user_id_of_item = int(user_id_str)
                                confirming_owner_name = _norm_owner_name(update.effective_user.username)
                                source_kind = item.get("source_kind")

                                await _increment_user_confirmation_count(user_id_of_item)
                                await _increment_owner_performance(confirming_owner_name, source_kind)
                                await _log_event("app_id", "confirmed", update, stored_app_id_raw, owner=confirming_owner_name)
                                await _clear_one_issued(user_id_of_item, "app_id", stored_app_id_raw)
                                log.info(f"Owner {confirming_owner_name} confirmed App ID {stored_app_id_raw}. Counted and cleared for user {user_id_of_item}")
                                found_and_counted = True
                                break

                    if not found_and_counted:
                        suggestion = _find_closest_app_id(app_id_confirmed_raw)
                        if suggestion:
                            reply_text = (
                                f"Wrong ID. Did you mean <code>{suggestion}</code>?\n\n"
                                f"Tap to copy and send again:\n"
                                f"<code>+1 {suggestion}</code>"
                            )
                            await msg.reply_html(reply_text)
                        else:
                            await msg.reply_text("Wrong ID, please check.")
                        log.warning(f"Received confirmation for incorrect App ID '{app_id_confirmed_raw}' from {update.effective_user.username}.")
            return

        # ================================================================
        # START OF MODIFIED BLOCK FOR CLEARING_GROUP_ID
        # ================================================================
        elif chat_id == CLEARING_GROUP_ID:
            # Find any pending items mentioned in the message
            pending_usernames = {item['value'] for item in _issued_bucket("username").get(str(uid), [])}
            pending_whatsapps = {item['value'] for item in _issued_bucket("whatsapp").get(str(uid), [])}
            found_usernames = {f"@{u}" for u in EXTRACT_USERNAMES_RX.findall(text)}
            found_phones = EXTRACT_PHONES_RX.findall(text)

            values_found_in_message = set()
            for u in found_usernames:
                if u in pending_usernames:
                    values_found_in_message.add(u)
            for p in found_phones:
                for pending_p in pending_whatsapps:
                    if _norm_phone(p) == _norm_phone(pending_p):
                        values_found_in_message.add(pending_p)


            # --- START CHANGE 1: Find app_id_match *before* validation ---
            # Find a new App ID in the message
            app_id_match = APP_ID_RX.search(text)

            # An actionable message must have a pending item or a new app ID
            if not values_found_in_message and not app_id_match:
                return
            # --- END CHANGE 1 ---

            # --- START CHANGE 2: Un-indent validation block ---
            # Now validation runs if *either* a pending item OR an app_id is found.
            found_country, country_status = _find_country_in_text(text)
            age = _find_age_in_text(text)
            is_allowed = True
            rejection_reason = ""

            # Tiered validation checks
            if country_status == 'not_allowed':
                is_allowed = False
                rejection_reason = f"Country '{found_country}' is not on the allowed list."

            # MODIFICATION: Added new validation logic here
            elif country_status: # If it's a known country, check bans and limits

                # 1. Check for manual (permanent) bans
                user_manual_bans = USER_COUNTRY_BANS.get(uid, set())
                if country_status in user_manual_bans:
                    is_allowed = False
                    rejection_reason = f"You are manually banned from submitting for the country '{found_country}'."
                else:
                    # 2. Check for automatic (daily) limit
                    current_country_count = await _get_user_country_count(uid, country_status)
                    if current_country_count >= COUNTRY_DAILY_LIMIT:
                        is_allowed = False
                        rejection_reason = f"You have reached the daily limit ({COUNTRY_DAILY_LIMIT}) for '{found_country}'. Please try again tomorrow."
                    else:
                        # 3. Check for India age limit
                        if country_status == 'india' and (age is None or age < 30):
                            is_allowed = False
                            rejection_reason = f"Age must be provided and must be 30 or older for India (found: {age})."

            if not is_allowed:
                # --- START CHANGE 3: Modified rejection message ---
                # This now handles cases where no pending item was found, but an app_id was.
                item_for_reply = "an item"
                if values_found_in_message:
                    first_offending_value = next(iter(values_found_in_message))
                    item_type = "username" if first_offending_value.startswith('@') else "whatsapp"
                    item_for_reply = f"that {item_type} (<code>{first_offending_value}</code>)"
                elif app_id_match:
                    app_id_val = f"@{app_id_match.group(2)}"
                    item_for_reply = f"that App ID (<code>{app_id_val}</code>)"

                reply_text = (f"{mention_user_html(uid)}, your submission was rejected. Reason: {rejection_reason}\n"
                              f"Please use {item_for_reply} for another customer.")
                # --- END CHANGE 3 ---
                await msg.reply_html(reply_text)
                log.warning(f"Rejected post from user {uid}. Reason: {rejection_reason}. No items cleared.")
                return
            else:
                # --- 2. Validation Passed. Process the message. ---
                if country_status:
                    await _increment_user_country_count(uid, country_status)
                    log.info(f"Incremented country count for user {uid} for '{country_status}'")

                # --- START CHANGE 4: Move forwarding block ---
                # This block is now un-indented and will run if validation passes
                # for *any* actionable message.
                # ========================================================
                log.info(f"Validation passed. Forwarding message {msg.message_id} using forward_message.")
                try:
                    await context.bot.forward_message(
                        chat_id=FORWARD_GROUP_ID, # The new target group ID
                        from_chat_id=chat_id,
                        message_id=msg.message_id
                    )
                    log.info(f"Forwarded successfully cleared message {msg.message_id} to {FORWARD_GROUP_ID}.")
                except Exception as e:
                    log.error(f"Failed to forward cleared message {msg.message_id} to {FORWARD_GROUP_ID}: {e}")
                # ========================================================
                # --- END CHANGE 4 ---
            # --- END CHANGE 2 ---

            # --- 4. Process App ID and clear source item ---
            # This logic was already here, but now it runs *after* validation and forwarding.
            # It is no longer inside an `else` block.
            if app_id_match:
                app_id = f"@{app_id_match.group(2)}"
                source_item_to_clear, source_kind = None, None

                # Priority 1: Link to an item explicitly mentioned in this message
                if values_found_in_message:
                    value = next(iter(values_found_in_message))
                    kind = "whatsapp" if _looks_like_phone(value) else "username"
                    for item in _issued_bucket(kind).get(str(uid), []):
                        if item.get("value") == value:
                            source_item_to_clear, source_kind = item, kind
                            break
                # Priority 2 (Fallback): Link to the most recently issued item for this user
                else:
                    last_item, last_ts = None, datetime.min.replace(tzinfo=TIMEZONE)
                    for kind in ("username", "whatsapp"):
                        user_items = _issued_bucket(kind).get(str(uid), [])
                        if user_items:
                            latest_in_kind = user_items[-1]
                            item_ts = datetime.fromisoformat(latest_in_kind["ts"])
                            if item_ts > last_ts:
                                last_ts = item_ts
                                last_item = latest_in_kind
                                last_item['kind'] = kind # Store kind for later
                    if last_item:
                        source_item_to_clear, source_kind = last_item, last_item['kind']


                # If we found a source item, log the new App ID and clear the source
                if source_item_to_clear and source_kind:
                    context_data = {
                        "source_owner": source_item_to_clear.get("owner"),
                        "source_kind": source_kind
                    }
                    value_to_clear = source_item_to_clear.get("value")
                    await _set_issued(uid, chat_id, "app_id", app_id, context_data=context_data)
                    await _log_event("app_id", "issued", update, app_id, owner=context_data.get("source_owner", ""))
                    log.info(f"Recorded App ID '{app_id}' for user {uid}, linked to {source_kind} '{value_to_clear}'")
                    if await _clear_one_issued(uid, source_kind, value_to_clear):
                        await _log_event(source_kind, "cleared", update, value_to_clear)
                        log.info(f"Auto-cleared pending {source_kind} for user {uid}: {value_to_clear}")

                else:
                    # Treat unlinked App IDs as valid entries, storing them for later confirmation
                    context_data = {"source_owner": "unknown", "source_kind": "app_id"}
                    await _set_issued(uid, chat_id, "app_id", app_id, context_data=context_data)
                    await _log_event("app_id", "issued", update, app_id)
                    log.info(
                        f"Recorded App ID '{app_id}' for user {uid} without a source item"
                    )
            return # End of processing for this group
        # ================================================================
        # END OF MODIFIED BLOCK
        # ================================================================

        elif chat_id == REQUEST_GROUP_ID:
            if uid not in WHITELISTED_USERS and not _is_admin(update.effective_user):
                return

            if NEED_USERNAME_RX.match(text):
                # Cooldown check for usernames
                now = datetime.now(TIMEZONE)
                last_req_ts_str = state.setdefault("username_last_request_ts", {}).get(str(uid))
                if last_req_ts_str:
                    last_req_ts = datetime.fromisoformat(last_req_ts_str)
                    if (now - last_req_ts) < timedelta(minutes=1):
                        await msg.reply_text("អ្នកអាចស្នើសុំ username បានតែម្តងគត់ក្នុងមួយនាទី។ សូមរង់ចាំ។")
                        return

                rec = await _next_from_username_pool()
                reply = "No available username." if not rec else f"@{rec['owner']}\n{rec['username']}"
                await msg.reply_text(reply)
                if rec:
                    # Record the time of this successful request
                    state.setdefault("username_last_request_ts", {})[str(uid)] = now.isoformat()
                    await _set_issued(uid, chat_id, "username", rec["username"], context_data={"owner": rec["owner"]})
                    await _log_event("username", "issued", update, rec["username"], owner=rec["owner"])
                    await _increment_user_activity(uid, "username")
                return

            if NEED_WHATSAPP_RX.match(text):
                if uid in WHATSAPP_BANNED_USERS:
                    await msg.reply_text("អ្នកត្រូវបានហាមឃាត់ជាអចិន្ត្រៃយ៍ពីការស្នើសុំលេខ WhatsApp ។")
                    return

                # Check for temporary bans
                temp_bans = state.get("whatsapp_temp_bans", {})
                if str(uid) in temp_bans:
                    ban_expires_ts = datetime.fromisoformat(temp_bans[str(uid)])
                    now = datetime.now(TIMEZONE)

                    if now < ban_expires_ts:
                        remaining_time = ban_expires_ts - now
                        minutes_left = round(remaining_time.total_seconds() / 60)
                        await msg.reply_text(f"អ្នកត្រូវបានហាមឃាត់ជាបណ្ដោះអាសន្ន។ សូមព្យាយាមម្តងទៀតក្នុងរយៈពេល {minutes_left} នាទីទៀត។")
                        return
                    else:
                        # Ban has expired, remove it
                        del temp_bans[str(uid)]
                        await save_state()
                        log.info(f"Temporary WhatsApp ban for user {uid} has expired and been removed.")

                # Cooldown check
                now = datetime.now(TIMEZONE)
                last_req_ts_str = state.setdefault("whatsapp_last_request_ts", {}).get(str(uid))
                if last_req_ts_str:
                    last_req_ts = datetime.fromisoformat(last_req_ts_str)
                    if (now - last_req_ts) < timedelta(minutes=3):
                        await msg.reply_text("អ្នកអាចស្នើសុំលេខ WhatsApp បានតែម្តងគត់ក្នុងរយៈពេល 3 នាទី។ សូមរង់ចាំ។")
                        return

                username_count, whatsapp_count = await _get_user_activity(uid)
                has_bonus = username_count > USERNAME_THRESHOLD_FOR_BONUS

                if not has_bonus and whatsapp_count >= USER_WHATSAPP_LIMIT:
                    await msg.reply_text(f"អ្នកបានស្នើសុំ WhatsApp គ្រប់ចំនួនកំណត់សម្រាប់ថ្ងៃនេះហើយ។\nសូមស្នើសុំ username ឱ្យលើសពី {USERNAME_THRESHOLD_FOR_BONUS} ដើម្បីទទួលបានការស្នើសុំ WhatsApp បន្ថែមទៀតដោយគ្មានដែនកំណត់។")
                    return

                rec = await _next_from_whatsapp_pool()
                reply = "No available WhatsApp."
                if rec:
                    if await _wa_quota_reached(rec["number"]):
                        reply = "No available WhatsApp (daily limit may be reached)."
                        rec = None
                    else:
                        reply = f"@{rec['owner']}\n{rec['number']}"

                await msg.reply_text(reply)
                if rec:
                    await _wa_inc_count(_norm_phone(rec["number"]), _logical_day_today())
                    # Record the time of this successful request before saving state
                    state.setdefault("whatsapp_last_request_ts", {})[str(uid)] = now.isoformat()
                    await _set_issued(uid, chat_id, "whatsapp", rec["number"], context_data={"owner": rec["owner"]})
                    await _log_event("whatsapp", "issued", update, rec["number"], owner=rec["owner"])
                    await _increment_user_activity(uid, "whatsapp")
                return

        m_owner = WHO_USING_REGEX.match(text)
        if m_owner:
            handle, phone = m_owner.groups()
            if handle:
                key = _norm_handle(handle); hits = HANDLE_INDEX.get(key, [])
                owners = sorted({h['owner'] for h in hits}) if hits else []
                reply = f"Owner of username @{key} → " + (", ".join(f"@{o}" for o in owners) if owners else "not found")
            else:
                pnorm = _norm_phone(phone); rec = PHONE_INDEX.get(pnorm)
                if rec and rec.get("channel") == "whatsapp": reply = f"Owner of WhatsApp {phone} → @{rec['owner']}"
                elif rec: reply = f"Owner of number {phone} → @{rec['owner']} (@{rec.get('telegram') or '-'})"
                else: reply = f"Owner of number {phone} → not found"

            await msg.reply_text(reply)
            return

# =============================
# MAIN
# =============================
async def post_initialization(application: Application):
    """Runs once after the bot is initialized."""
    await get_db_pool()
    await setup_database()
    await load_state()
    await _migrate_state_if_needed()
    await load_owner_directory()
    await load_whatsapp_bans()
    await load_user_country_bans() # NEW: Load country bans
    await load_admins()
    await load_whitelisted_users()

async def post_shutdown(application: Application):
    """Runs once before the bot shuts down."""
    await close_db_pool()


if __name__ == "__main__":
    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .post_init(post_initialization)
        .post_shutdown(post_shutdown)
        .build()
    )

    if app.job_queue:
        app.job_queue.run_repeating(check_reminders, interval=60, first=60)
        # MODIFIED: Changed the check interval to 60 minutes (3600s).
        app.job_queue.run_repeating(check_request_ratio_and_stop_whatsapp, interval=3600, first=3600)
        # NEW JOB for clearing expired IDs, runs every hour
        app.job_queue.run_repeating(_clear_expired_app_ids, interval=3600, first=3600)
        reset_time = time(hour=5, minute=31, tzinfo=TIMEZONE)
        app.job_queue.run_daily(daily_reset, time=reset_time)

    app.add_handler(MessageHandler(filters.ALL & ~filters.StatusUpdate.ALL, on_message))

    log.info("Bot is starting...")
    app.run_polling(drop_pending_updates=True, allowed_updates=Update.ALL_TYPES)



