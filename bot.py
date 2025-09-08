#!/usr/bin/env python3
import os
import json
import csv
import asyncio
import logging
from datetime import datetime, timedelta, date
from typing import Dict, Optional, List, Tuple
import re
import time

import pytz
import psycopg2
import psycopg2.extras
from telegram import Update
from telegram.constants import ChatType, ParseMode
from telegram.ext import Application, ContextTypes, MessageHandler, filters

# = a===========================
# LOGGING
# =============================
logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO,
)
log = logging.getLogger("bot")

# =============================
# CONFIG
# =============================
def get_env_variable(var_name: str) -> str:
    """Gets an environment variable or raises an error."""
    value = os.getenv(var_name)
    if not value:
        raise RuntimeError(
            f"Missing required environment variable '{var_name}'. "
            "Please set it in your hosting environment."
        )
    return value

BOT_TOKEN = get_env_variable("BOT_TOKEN")
DATABASE_URL = get_env_variable("DATABASE_URL")
ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "excelmerge") # telegram username (without @)

TIMEZONE = pytz.timezone("Asia/Phnom_Penh")

# =============================
# DATABASE SETUP & HELPERS
# =============================
db_lock = asyncio.Lock()

def get_db_connection():
    """Establishes a connection to the PostgreSQL database."""
    try:
        conn = psycopg2.connect(DATABASE_URL)
        return conn
    except psycopg2.OperationalError as e:
        log.error("Could not connect to database: %s", e)
        raise

def setup_database():
    """Ensures the required tables exist in the database."""
    log.info("Setting up database schema...")
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            # Table for key-value JSON storage (for state and owners)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS kv_storage (
                    key TEXT PRIMARY KEY,
                    data JSONB NOT NULL,
                    updated_at TIMESTAMPTZ DEFAULT NOW()
                );
            """)
            # Table for the audit log (replaces the CSV)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS audit_log (
                    id SERIAL PRIMARY KEY,
                    ts_local TIMESTAMPTZ NOT NULL,
                    chat_id BIGINT,
                    message_id BIGINT,
                    user_id BIGINT,
                    user_first TEXT,
                    user_username TEXT,
                    kind TEXT,
                    action TEXT,
                    value TEXT,
                    owner TEXT
                );
            """)
            conn.commit()
        log.info("Database schema is ready.")
    finally:
        conn.close()

# =============================
# STATE (Now in Database)
# =============================
BASE_STATE = {
    "user_names": {},
    "rr": {
        "username_owner_idx": 0, "username_entry_idx": {},
        "wa_owner_idx": 0, "wa_entry_idx": {},
    },
    "issued": {"username": {}, "whatsapp": {}},
}
state: Dict = {k: (v.copy() if isinstance(v, dict) else v) for k, v in BASE_STATE.items()}

def load_state():
    """Loads state from the 'state' key in the database."""
    global state
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT data FROM kv_storage WHERE key = 'state'")
            result = cur.fetchone()
            if result:
                loaded = result[0]
                state = {**BASE_STATE, **loaded}
                log.info("Bot state loaded from database.")
            else:
                log.info("No state found in database, using base state.")
                state = {k: (v.copy() if isinstance(v, dict) else v) for k, v in BASE_STATE.items()}
                save_state() # Save initial state
    except Exception as e:
        log.warning("Failed to load state from DB: %s", e)
    finally:
        conn.close()

    # Ensure nested dictionaries exist
    state.setdefault("rr", {}).setdefault("username_entry_idx", {})
    state["rr"].setdefault("wa_entry_idx", {})
    state.setdefault("issued", {}).setdefault("username", {})
    state["issued"].setdefault("whatsapp", {})


def save_state():
    """Saves the current state to the 'state' key in the database."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO kv_storage (key, data)
                VALUES ('state', %s)
                ON CONFLICT (key) DO UPDATE
                SET data = EXCLUDED.data, updated_at = NOW();
            """, (json.dumps(state),))
            conn.commit()
    except Exception as e:
        log.warning("Failed to save state to DB: %s", e)
    finally:
        conn.close()

# =============================
# OWNER DIRECTORY (Now in Database)
# =============================
OWNER_DATA: List[dict] = []
HANDLE_INDEX: Dict[str, List[dict]] = {}
PHONE_INDEX: Dict[str, dict] = {}
USERNAME_POOL: List[Dict] = []
WHATSAPP_POOL: List[Dict] = []

# ... (Helper functions like _norm_handle, _is_admin, etc. remain the same) ...
def _norm_handle(h: str) -> str: return re.sub(r"^@", "", (h or "").strip().lower())
def _norm_phone(p: str) -> str: return re.sub(r"\D+", "", (p or ""))
def _norm_owner_name(s: str) -> str:
    s = (s or "").strip()
    if s.startswith("@"): s = s[1:]
    return s.lower()
def _is_admin(update: Update) -> bool:
    u = update.effective_user
    if not u: return False
    return (u.username or "").lower() == ADMIN_USERNAME.lower()
def _owner_is_paused(group: dict) -> bool:
    if group.get("disabled") or group.get("active") is False: return True
    until = group.get("disabled_until")
    if until:
        try:
            dt = datetime.fromisoformat(until)
            if dt.tzinfo is None: dt = TIMEZONE.localize(dt)
            return datetime.now(tz=TIMEZONE) < dt
        except Exception: return False
    return False

def _ensure_owner_shape(g: dict) -> dict:
    g.setdefault("owner", "")
    g.setdefault("disabled", False)
    g.setdefault("entries", [])
    g.setdefault("whatsapp", [])
    norm_entries = []
    for e in g.get("entries", []):
        if isinstance(e, dict):
            e.setdefault("telegram", ""); e.setdefault("phone", ""); e.setdefault("disabled", False)
            norm_entries.append(e)
    g["entries"] = norm_entries
    norm_wa = []
    for w in g.get("whatsapp", []):
        if isinstance(w, dict):
            w.setdefault("number", w.get("number") or w.get("phone") or "")
            w.setdefault("disabled", False)
            if (w["number"] or "").strip(): norm_wa.append(w)
        elif isinstance(w, str) and w.strip():
            norm_wa.append({"number": w.strip(), "disabled": False})
    g["whatsapp"] = norm_wa
    return g

def save_owner_directory():
    """Saves the OWNER_DATA list to the 'owners' key in the database."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO kv_storage (key, data)
                VALUES ('owners', %s)
                ON CONFLICT (key) DO UPDATE
                SET data = EXCLUDED.data, updated_at = NOW();
            """, (json.dumps(OWNER_DATA),))
            conn.commit()
    except Exception as e:
        log.error("Failed to save owner directory to DB: %s", e)
    finally:
        conn.close()

def load_owner_directory():
    """Loads owner data from DB and rebuilds in-memory pools and indexes."""
    global OWNER_DATA, HANDLE_INDEX, PHONE_INDEX, USERNAME_POOL, WHATSAPP_POOL
    OWNER_DATA, HANDLE_INDEX, PHONE_INDEX = [], {}, {}
    USERNAME_POOL, WHATSAPP_POOL = [], []

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT data FROM kv_storage WHERE key = 'owners'")
            result = cur.fetchone()
            if result and result[0]:
                OWNER_DATA = result[0]
            else:
                log.warning("Owner directory not found in database. Will create on first add.")
                OWNER_DATA = []
    except Exception as e:
        log.error("Failed to load owners from DB: %s", e)
        OWNER_DATA = []
    finally:
        conn.close()

    OWNER_DATA = [_ensure_owner_shape(dict(g)) for g in OWNER_DATA]

    # Rebuilding logic remains the same, operating on the loaded OWNER_DATA
    for group in OWNER_DATA:
        owner = _norm_owner_name(group.get("owner") or "")
        if not owner or _owner_is_paused(group): continue
        usernames: List[str] = []
        for entry in group.get("entries", []):
            if entry.get("disabled"): continue
            tel = (entry.get("telegram") or "").strip()
            ph  = (entry.get("phone") or "").strip()
            if tel:
                handle_shown = tel if tel.startswith("@") else f"@{tel}"
                usernames.append(handle_shown)
                HANDLE_INDEX.setdefault(_norm_handle(tel), []).append(
                    {"owner": owner, "phone": ph, "telegram": tel, "channel": "telegram"})
            if ph:
                PHONE_INDEX[_norm_phone(ph)] = {"owner": owner, "phone": ph, "telegram": tel, "channel": "telegram"}
        if usernames: USERNAME_POOL.append({"owner": owner, "usernames": usernames})
        numbers: List[str] = []
        for w in group.get("whatsapp", []):
            if w.get("disabled"): continue
            num = (w.get("number") or "").strip()
            if num:
                numbers.append(num)
                PHONE_INDEX[_norm_phone(num)] = {"owner": owner, "phone": num, "telegram": None, "channel": "whatsapp"}
        if numbers: WHATSAPP_POOL.append({"owner": owner, "numbers": numbers})
    log.info("[owner_directory] owners(active): usernames=%d, whatsapp=%d | handles=%d, phones=%d",
             len(USERNAME_POOL), len(WHATSAPP_POOL), len(HANDLE_INDEX), len(PHONE_INDEX))

# =============================
# Rotation-preserving rebuild & Round-robin helpers
# (These functions do not need changes as they operate on in-memory variables)
# =============================
def _owner_list_from_pool(pool) -> List[str]: return [blk["owner"] for blk in pool]
def _preserve_owner_pointer(old_list: List[str], new_list: List[str], old_idx: int) -> int:
    if not new_list: return 0
    old_list = list(old_list or []);
    if not old_list: return 0
    old_idx = old_idx % len(old_list)
    start_owner = old_list[old_idx]
    if start_owner in new_list: return new_list.index(start_owner)
    n = len(old_list)
    for step in range(1, n + 1):
        cand = old_list[(old_idx + step) % n]
        if cand in new_list: return new_list.index(cand)
    return 0

def _preserve_entry_indices(rr_map: Dict[str, int], new_pool: List[Dict], list_key: str):
    valid_owners = {blk["owner"]: len(blk.get(list_key, []) or []) for blk in new_pool}
    for owner in list(rr_map.keys()):
        if owner not in valid_owners: rr_map.pop(owner, None)
    for owner, sz in valid_owners.items():
        if sz <= 0: rr_map.pop(owner, None)
        else: rr_map[owner] = rr_map.get(owner, 0) % sz

def _rebuild_pools_preserving_rotation():
    asyncio.run(db_lock.acquire())
    try:
        old_user_owner_list = _owner_list_from_pool(USERNAME_POOL)
        old_wa_owner_list   = _owner_list_from_pool(WHATSAPP_POOL)
        rr = state.setdefault("rr", {})
        old_user_owner_idx = rr.get("username_owner_idx", 0)
        old_wa_owner_idx   = rr.get("wa_owner_idx", 0)
        old_user_entry_idx = dict(rr.get("username_entry_idx", {}))
        old_wa_entry_idx   = dict(rr.get("wa_entry_idx", {}))
        load_owner_directory()
        save_owner_directory()
        new_user_owner_list = _owner_list_from_pool(USERNAME_POOL)
        new_wa_owner_list   = _owner_list_from_pool(WHATSAPP_POOL)
        rr["username_owner_idx"] = _preserve_owner_pointer(old_user_owner_list, new_user_owner_list, old_user_owner_idx)
        rr["wa_owner_idx"] = _preserve_owner_pointer(old_wa_owner_list, new_wa_owner_list, old_wa_owner_idx)
        rr.setdefault("username_entry_idx", old_user_entry_idx)
        rr.setdefault("wa_entry_idx", old_wa_entry_idx)
        _preserve_entry_indices(rr["username_entry_idx"], USERNAME_POOL, "usernames")
        _preserve_entry_indices(rr["wa_entry_idx"], WHATSAPP_POOL, "numbers")
        save_state()
    finally:
        db_lock.release()

def _next_from_username_pool() -> Optional[Dict[str, str]]:
    # ... Function implementation is unchanged ...
    if not USERNAME_POOL: return None
    rr = state["rr"]
    idx = rr.get("username_owner_idx", 0) % len(USERNAME_POOL)
    for _ in range(len(USERNAME_POOL)):
        block = USERNAME_POOL[idx]
        arr = block.get("usernames", [])
        if arr:
            ei = rr["username_entry_idx"].get(block["owner"], 0) % len(arr)
            result = {"owner": block["owner"], "username": arr[ei]}
            rr["username_entry_idx"][block["owner"]] = (ei + 1) % len(arr)
            rr["username_owner_idx"] = (idx + 1) % len(USERNAME_POOL)
            save_state()
            return result
        idx = (idx + 1) % len(USERNAME_POOL)
    return None

def _next_from_whatsapp_pool() -> Optional[Dict[str, str]]:
    # ... Function implementation is unchanged ...
    if not WHATSAPP_POOL: return None
    rr = state["rr"]
    owner_idx = rr.get("wa_owner_idx", 0) % len(WHATSAPP_POOL)
    for _ in range(len(WHATSAPP_POOL)):
        block = WHATSAPP_POOL[owner_idx]
        owner = block["owner"]
        numbers = block.get("numbers", [])
        if numbers:
            ni = rr["wa_entry_idx"].get(owner, 0) % len(numbers)
            number = numbers[ni]
            rr["wa_entry_idx"][owner] = (ni + 1) % len(numbers)
            rr["wa_owner_idx"] = (owner_idx + 1) % len(WHATSAPP_POOL)
            save_state()
            return {"owner": owner, "number": number}
        owner_idx = (owner_idx + 1) % len(WHATSAPP_POOL)
    return None

# =============================
# REGEXES (unchanged)
# =============================
WHO_USING_REGEX = re.compile(r"^\s*who(?:['\u2019]s| is)\s+using\s+(?:@?([A-Za-z0-9_\.]+)|(\+?\d[\d\s\-]{6,}\d))\s*$", re.IGNORECASE)
NEED_USERNAME_RX = re.compile(r"^\s*i\s*need\s*(?:user\s*name|username)\s*$", re.IGNORECASE)
NEED_WHATSAPP_RX = re.compile(r"^\s*i\s*need\s*(?:id\s*)?whats?app\s*$", re.IGNORECASE)
STOP_OPEN_RX = re.compile(r"^\s*(stop|open)\s+@?(.+?)\s*$", re.IGNORECASE)
ADD_OWNER_RX = re.compile(r"^\s*add\s+owner\s+@?(.+?)\s*$", re.IGNORECASE)
ADD_USERNAME_RX = re.compile(r"^\s*add\s+username\s+@([A-Za-z0-9_]{3,})\s+to\s+@?(.+?)\s*$", re.IGNORECASE)
ADD_WHATSAPP_RX = re.compile(r"^\s*add\s+whats?app\s+(\+?\d[\d\s\-]{6,}\d)\s+to\s+@?(.+?)\s*$", re.IGNORECASE)
DEL_USERNAME_RX = re.compile(r"^\s*delete\s+username\s+@([A-Za-z0-9_]{3,})\s*$", re.IGNORECASE)
DEL_WHATSAPP_RX = re.compile(r"^\s*delete\s+whats?app\s+(\+?\d[\d\s\-]{6,}\d)\s*$", re.IGNORECASE)
LIST_OWNERS_RX = re.compile(r"^\s*list\s+owners\s*$", re.IGNORECASE)
LIST_OWNER_DETAIL_RX = re.compile(r"^\s*list\s+owner\s+@?(.+?)\s*$", re.IGNORECASE)
LIST_DISABLED_RX = re.compile(r"^\s*list\s+disabled\s*$", re.IGNORECASE)
SEND_REPORT_RX = re.compile(r"^\s*(?:send\s+report|report)(?:\s+(yesterday|today|\d{4}-\d{2}-\d{2}))?\s*$", re.IGNORECASE)

# =============================
# CSV AUDIT LOG (Now in Database)
# =============================
def _log_event(kind: str, action: str, update: Update, value: str, owner: str = ""):
    """Writes an audit event to the audit_log table."""
    conn = get_db_connection()
    try:
        u = update.effective_user
        m = update.effective_message
        c = update.effective_chat
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO audit_log (
                    ts_local, chat_id, message_id, user_id, user_first,
                    user_username, kind, action, value, owner
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                datetime.now(TIMEZONE), c.id if c else None, m.message_id if m else None,
                u.id if u else None, u.first_name if u else None, u.username if u else None,
                kind, action, value, owner or ""
            ))
            conn.commit()
    except Exception as e:
        log.warning("Log write to DB failed: %s", e)
    finally:
        conn.close()

# =============================
# UTIL (unchanged, but state helpers now talk to DB indirectly)
# =============================
def cache_user_info(user): state.setdefault("user_names", {})[str(user.id)] = {"first_name": user.first_name or "", "username": user.username or ""}
def mention_user_html(user_id: int) -> str:
    info = state.setdefault("user_names", {}).get(str(user_id), {})
    name = info.get("first_name") or info.get("username") or str(user_id)
    return f'<a href="tg://user?id={user_id}">{name}</a>'
def _issued_bucket(kind: str) -> Dict[str, dict]: return state.setdefault("issued", {}).setdefault(kind, {})
def _set_issued(user_id: int, kind: str, value: str):
    _issued_bucket(kind)[str(user_id)] = {"value": value, "ts": datetime.now(TIMEZONE).isoformat()}
    save_state()
def _get_issued(user_id: int, kind: str) -> Optional[str]: return _issued_bucket(kind).get(str(user_id), {}).get("value")
def _clear_issued(user_id: int, kind: str):
    _issued_bucket(kind).pop(str(user_id), None)
    save_state()
def _value_in_text(value: Optional[str], text: str) -> bool:
    if not value: return False
    v = value.strip()
    if v.startswith("@"): return v.lower() in (text or "").lower()
    def norm(s): return re.sub(r"\D+", "", s or "")
    return norm(v) and (norm(v) in norm(text or ""))
async def _deny_for_unreturned(msg, user_id: int, kind: str):
    pending = _get_issued(user_id, kind)
    label = "username" if kind == "username" else "WhatsApp"
    if pending:
        await msg.chat.send_message(
            f"{mention_user_html(user_id)} please provide your previous {label} first: {pending}\n"
            f"Just include it in your next profile message and I’ll accept it automatically.",
            reply_to_message_id=msg.message_id, parse_mode=ParseMode.HTML,
        )

# =============================
# EXCEL DAILY REPORT (Reads from DB)
# =============================
def _logical_day_of(ts: datetime) -> date:
    """Takes a timezone-aware datetime object."""
    shifted = ts.astimezone(TIMEZONE) - timedelta(hours=5) # 05:00 boundary
    return shifted.date()

def _read_log_rows() -> List[dict]:
    """Reads all rows from the audit_log table."""
    conn = get_db_connection()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute("SELECT * FROM audit_log ORDER BY ts_local")
            return [dict(row) for row in cur.fetchall()]
    except Exception as e:
        log.error("Failed to read log rows from DB: %s", e)
        return []
    finally:
        conn.close()

# ... (_compute_daily_summary, _style_and_save_excel, _parse_report_day remain mostly the same)
def _compute_daily_summary(target_day: date) -> List[dict]:
    rows = _read_log_rows()
    day_rows = [r for r in rows if _logical_day_of(r["ts_local"]) == target_day]
    users: Dict[str, dict] = {}
    for r in day_rows:
        user_key = r.get("user_first") or r.get("user_username") or str(r.get("user_id"))
        d = users.setdefault(user_key, {"username_issued": [], "username_cleared": [], "wa_issued": [], "wa_cleared": []})
        kind, action, value, owner = r["kind"], r["action"], r["value"], r.get("owner","")
        if kind == "username":
            if action == "issued": d["username_issued"].append((value, owner))
            if action == "cleared": d["username_cleared"].append(value)
        elif kind == "whatsapp":
            if action == "issued": d["wa_issued"].append((value, owner))
            if action == "cleared": d["wa_cleared"].append(value)
    out = []
    for user, d in users.items():
        issued_user_pairs = d["username_issued"]; issued_wa_pairs = d["wa_issued"]
        issued_user_vals  = [v for v, _ in issued_user_pairs]; issued_wa_vals = [v for v, _ in issued_wa_pairs]
        notback_user = [v for v in issued_user_vals if v not in d["username_cleared"]]
        notback_wa = [v for v in issued_wa_vals if v not in d["wa_cleared"]]
        owners_user = sorted({own for (v, own) in issued_user_pairs if v in notback_user and own})
        owners_wa = sorted({own for (v, own) in issued_wa_pairs if v in notback_wa and own})
        out.append({
            "Day": target_day.isoformat(), "User": str(user),
            "Total username receive": len(issued_user_vals), "Total whatsapp receive": len(issued_wa_vals),
            "Total username provide back": len(d["username_cleared"]), "Total whatsapp provide back": len(d["wa_cleared"]),
            "Username not provide back": ", ".join(notback_user), "Owner of username": ", ".join(owners_user),
            "Whatsapp not provide back": ", ".join(notback_wa), "Owner of whatsapp": ", ".join(owners_wa),
        })
    out.sort(key=lambda r: r["User"].lower())
    return out
def _style_and_save_excel(rows: List[dict], out_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed. Please add it to requirements.txt")
    headers = ["Day","User","Total username receive","Total whatsapp receive","Total username provide back","Total whatsapp provide back", "Username not provide back","Owner of username","Whatsapp not provide back","Owner of whatsapp"]
    wb = Workbook(); ws = wb.active; ws.title = "Summary"; ws.append(headers)
    for r in rows: ws.append([r.get(h, "") for h in headers])
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = border
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
        fill = PatternFill(start_color="F2F2F2" if i % 2 == 0 else "FFFFFF", fill_type="solid")
        for cell in row: cell.alignment = center; cell.border = border; cell.fill = fill
    for col in ws.columns:
        max_len = 0; letter = col[0].column_letter
        for c in col:
            if c.value: max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)
    wb.save(out_path)
def _send_daily_excel(update: Update, msg, target_day: date):
    rows = _compute_daily_summary(target_day)
    if not rows: return ("No data for that day.", None)
    fname = f"daily_summary_{target_day.isoformat()}.xlsx"
    path = os.path.join(os.getcwd(), fname) # Temp storage is fine here
    _style_and_save_excel(rows, path)
    return (None, path)
def _parse_report_day(arg: Optional[str]) -> date:
    now = datetime.now(TIMEZONE)
    if not arg or arg.lower() == "today": return (now - timedelta(hours=5)).date()
    if arg.lower() == "yesterday": return (now - timedelta(hours=5, days=1)).date()
    try: return datetime.strptime(arg, "%Y-%m-%d").date()
    except Exception: return (now - timedelta(hours=5)).date()

# =============================
# ADMIN COMMANDS (logic is unchanged, but calls _rebuild_pools)
# =============================
# ... The entire `_handle_admin_command` function and its helpers can be copied here ...
def _parse_duration(duration_str: str) -> Optional[timedelta]:
    m = re.match(r"(\d+)\s*(m|h|d|w)", (duration_str or "").strip().lower())
    if not m: return None
    val, unit = m.groups()
    val = int(val)
    if unit == 'm': return timedelta(minutes=val)
    if unit == 'h': return timedelta(hours=val)
    if unit == 'd': return timedelta(days=val)
    if unit == 'w': return timedelta(weeks=val)
    return None

def _find_owner_group(name: str) -> Optional[dict]:
    norm_name = _norm_owner_name(name)
    for group in OWNER_DATA:
        if _norm_owner_name(group["owner"]) == norm_name: return group
    return None

def _handle_admin_command(text: str) -> Optional[str]:
    m = ADD_OWNER_RX.match(text)
    if m:
        name = _norm_owner_name(m.group(1))
        if _find_owner_group(name): return f"Owner '{name}' already exists."
        OWNER_DATA.append(_ensure_owner_shape({"owner": name}))
        _rebuild_pools_preserving_rotation(); return f"Owner '{name}' added."
    m = ADD_USERNAME_RX.match(text)
    if m:
        handle, owner_name = m.groups()
        owner = _find_owner_group(owner_name)
        if not owner: return f"Owner '{owner_name}' not found."
        norm_h = _norm_handle(handle)
        if any(_norm_handle(e.get("telegram")) == norm_h for e in owner["entries"]): return f"@{handle} already exists for owner {owner_name}."
        owner["entries"].append({"telegram": handle, "phone": "", "disabled": False})
        _rebuild_pools_preserving_rotation(); return f"Added username @{handle} to {owner_name}."
    m = ADD_WHATSAPP_RX.match(text)
    if m:
        num, owner_name = m.groups()
        owner = _find_owner_group(owner_name)
        if not owner: return f"Owner '{owner_name}' not found."
        norm_n = _norm_phone(num)
        if any(_norm_phone(w.get("number")) == norm_n for w in owner["whatsapp"]): return f"Number {num} already exists for owner {owner_name}."
        owner["whatsapp"].append({"number": num, "disabled": False})
        _rebuild_pools_preserving_rotation(); return f"Added WhatsApp {num} to {owner_name}."
    m = DEL_USERNAME_RX.match(text)
    if m:
        handle = m.group(1); norm_h = _norm_handle(handle); found = False
        for owner in OWNER_DATA:
            original_len = len(owner["entries"])
            owner["entries"] = [e for e in owner["entries"] if _norm_handle(e.get("telegram")) != norm_h]
            if len(owner["entries"]) < original_len: found = True
        if not found: return f"Username @{handle} not found."
        _rebuild_pools_preserving_rotation(); return f"Deleted username @{handle} from all owners."
    m = DEL_WHATSAPP_RX.match(text)
    if m:
        num = m.group(1); norm_n = _norm_phone(num); found = False
        for owner in OWNER_DATA:
            original_len = len(owner["whatsapp"])
            owner["whatsapp"] = [w for w in owner["whatsapp"] if _norm_phone(w.get("number")) != norm_n]
            if len(owner["whatsapp"]) < original_len: found = True
        if not found: return f"WhatsApp number {num} not found."
        _rebuild_pools_preserving_rotation(); return f"Deleted WhatsApp number {num} from all owners."
    m = STOP_OPEN_RX.match(text)
    if m:
        action, target = m.groups(); parts = target.split(); owner_name = parts[0]
        duration_str = parts[1] if len(parts) > 1 else None
        owner = _find_owner_group(owner_name)
        if not owner: return f"Owner '{owner_name}' not found."
        if action.lower() == "open":
            owner["disabled"] = False; owner.pop("disabled_until", None)
            _rebuild_pools_preserving_rotation(); return f"Opened (resumed) owner {owner_name}."
        else:
            duration = _parse_duration(duration_str)
            if duration:
                until = datetime.now(TIMEZONE) + duration
                owner["disabled_until"] = until.isoformat(); owner["disabled"] = False
                _rebuild_pools_preserving_rotation(); return f"Stopped (paused) owner {owner_name} until {until.strftime('%Y-%m-%d %H:%M:%S')}."
            else:
                owner["disabled"] = True; owner.pop("disabled_until", None)
                _rebuild_pools_preserving_rotation(); return f"Stopped (paused) owner {owner_name} indefinitely."
    if LIST_OWNERS_RX.match(text):
        if not OWNER_DATA: return "No owners configured."
        lines = ["<b>Owner Roster:</b>"]
        for o in OWNER_DATA:
            status = "PAUSED" if _owner_is_paused(o) else "active"
            u_count = len([e for e in o.get("entries", []) if not e.get("disabled")])
            w_count = len([w for w in o.get("whatsapp", []) if not w.get("disabled")])
            lines.append(f"- <code>{o['owner']}</code> ({status}): {u_count} usernames, {w_count} whatsapps")
        return "\n".join(lines)
    if LIST_DISABLED_RX.match(text):
        disabled = [o for o in OWNER_DATA if _owner_is_paused(o)]
        if not disabled: return "No owners are currently disabled/paused."
        lines = ["<b>Disabled/Paused Owners:</b>"]
        for o in disabled:
            reason = f" (until {o['disabled_until']})" if o.get("disabled_until") else ""
            lines.append(f"- <code>{o['owner']}</code>{reason}")
        return "\n".join(lines)
    m = LIST_OWNER_DETAIL_RX.match(text)
    if m:
        owner = _find_owner_group(m.group(1))
        if not owner: return f"Owner '{m.group(1)}' not found."
        lines = [f"<b>Details for {owner['owner']}:</b>"]
        if owner.get("entries"): lines.extend(["<u>Usernames:</u>"] + [f"- @{e['telegram']}"] for e in owner["entries"])
        if owner.get("whatsapp"): lines.extend(["<u>WhatsApp Numbers:</u>"] + [f"- {w['number']}"] for w in owner["whatsapp"])
        if not owner.get("entries") and not owner.get("whatsapp"): lines.append("No entries found.")
        return "\n".join(lines)
    return None

# =============================
# MESSAGE HANDLER (Main logic loop)
# =============================
async def on_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_chat or not update.effective_user or \
       update.effective_chat.type not in (ChatType.GROUP, ChatType.SUPERGROUP):
        return

    msg = update.effective_message
    text = (msg.text or msg.caption or "").strip()
    uid = update.effective_user.id
    cache_user_info(update.effective_user)

    await db_lock.acquire()
    try:
        # Admin: Report Generation
        if _is_admin(update) and SEND_REPORT_RX.match(text):
            m = SEND_REPORT_RX.match(text)
            target_day = _parse_report_day(m.group(1))
            err, path = _send_daily_excel(update, msg, target_day)
            if err: await msg.chat.send_message(err, reply_to_message_id=msg.message_id)
            elif path:
                try:
                    await msg.chat.send_document(open(path, "rb"), filename=os.path.basename(path),
                        caption=f"Daily summary (logical day starting 05:00) — {target_day}",
                        reply_to_message_id=msg.message_id)
                except Exception as e:
                    await msg.chat.send_message(f"Failed to send report: {e}", reply_to_message_id=msg.message_id)
                finally:
                    if os.path.exists(path): os.remove(path)
            return

        # Admin: Console Commands
        if _is_admin(update):
            admin_reply = _handle_admin_command(text)
            if admin_reply:
                await msg.chat.send_message(admin_reply, reply_to_message_id=msg.message_id, parse_mode=ParseMode.HTML)
                return

        # Auto-clear holds
        for kind in ("username", "whatsapp"):
            pending = _get_issued(uid, kind)
            if pending and _value_in_text(pending, text):
                _clear_issued(uid, kind)
                _log_event(kind, "cleared", update, pending, owner="")
                log.info("Cleared pending %s for user %s.", kind, uid)

        # Feeders
        if NEED_USERNAME_RX.match(text):
            if _get_issued(uid, "username"): await _deny_for_unreturned(msg, uid, "username"); return
            rec = _next_from_username_pool()
            reply = "No available username." if not rec else f"@{rec['owner']}\n{rec['username']}"
            await msg.chat.send_message(reply, reply_to_message_id=msg.message_id)
            if rec: _set_issued(uid, "username", rec["username"]); _log_event("username", "issued", update, rec["username"], owner=rec["owner"])
            return

        if NEED_WHATSAPP_RX.match(text):
            if _get_issued(uid, "whatsapp"): await _deny_for_unreturned(msg, uid, "whatsapp"); return
            rec = _next_from_whatsapp_pool()
            reply = "No available WhatsApp." if not rec else f"@{rec['owner']}\n{rec['number']}"
            await msg.chat.send_message(reply, reply_to_message_id=msg.message_id)
            if rec: _set_issued(uid, "whatsapp", rec["number"]); _log_event("whatsapp", "issued", update, rec["number"], owner=rec["owner"])
            return

        # Lookups
        m_owner = WHO_USING_REGEX.match(text)
        if m_owner:
            handle, phone = m_owner.groups()
            if handle:
                key = _norm_handle(handle); hits = HANDLE_INDEX.get(key, [])
                owners = sorted({h['owner'] for h in hits}) if hits else []
                reply = f"Owner of username @{key} → " + (", ".join(f"@{o}" for o in owners) if owners else "not found")
            else:
                pnorm = _norm_phone(phone); rec = PHONE_INDEX.get(pnorm)
                if rec and rec.get("channel") == "whatsapp": reply = f"Owner of WhatsApp {phone} → @{rec['owner']}"
                elif rec: reply = f"Owner of number {phone} → @{rec['owner']} (@{rec.get('telegram') or '-'})"
                else: reply = f"Owner of number {phone} → not found"
            await msg.chat.send_message(reply, reply_to_message_id=msg.message_id)
            return

    finally:
        db_lock.release()

# =============================
# MAIN
# =============================
if __name__ == "__main__":
    setup_database()
    load_state()
    load_owner_directory()

    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(MessageHandler(filters.ALL & ~filters.StatusUpdate.ALL, on_message))

    log.info("Bot is starting...")
    app.run_polling(drop_pending_updates=True, allowed_updates=Update.ALL_TYPES)
